from flask import Blueprint, current_app, jsonify, render_template, request, send_file
from flask_login import current_user, login_required
import csv
import io
import os
import re
import traceback
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models import SiteContractMaster
from app.site_contract_master import VALID_SEGMENTS, ensure_contract_master_synced, manager_ids_match


subject_analysis_tool_bp = Blueprint("subject_analysis_tool", __name__, url_prefix="/tools/subject_analysis_tool")


_DECIMAL_INTEGER_PATTERN = re.compile(r"^\d+\.0+$")
_DIGITS_PATTERN = re.compile(r"^\d+$")


def normalize_contract_code(value):
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace(",", "").replace(" ", "").replace("\u3000", "")
    if _DECIMAL_INTEGER_PATTERN.match(text):
        text = text.split(".", 1)[0]
    if _DIGITS_PATTERN.match(text) and len(text) < 8:
        text = text.zfill(8)
    return text


def get_upload_folder():
    """一時アップロードフォルダを返す。"""
    folder = os.path.join(current_app.root_path, "static", "subject_analysis_tool", "uploads")
    os.makedirs(folder, exist_ok=True)
    return folder


def load_site_mapping_from_master():
    """現場リストPLUSの最新マスタからセグメント辞書、現場マスタ、警告を返す。"""
    ensure_contract_master_synced()

    rows = (
        SiteContractMaster.query
        .filter(SiteContractMaster.is_active.is_(True))
        .order_by(SiteContractMaster.contract_code.asc())
        .all()
    )

    site_dict = {}
    site_master = {}
    warnings = []
    matched_rows = 0

    for row in rows:
        matched_rows += 1
        contract_code = normalize_contract_code(row.contract_code)
        segment = str(row.segment or "").strip()
        if not contract_code:
            continue

        site_master[contract_code] = {
            "contract_code": contract_code,
            "segment": segment if segment in VALID_SEGMENTS else "",
            "site_manager_id": str(row.site_manager_id or "").strip(),
            "site_manager_name": str(row.site_manager_name or "").strip(),
            "site_name": str(row.site_name or "").strip(),
        }

        if segment not in VALID_SEGMENTS:
            warnings.append(f"segment missing: {contract_code} - {row.site_name}")
            continue
        site_dict[contract_code] = segment
    if matched_rows == 0:
        raise ValueError("現場リストPLUSの契約マスタがまだ作成されていません。現場リストPLUSを開いて同期してください")
    if matched_rows > 0 and not site_dict:
        raise ValueError("現場リストPLUSの現場はありますが、セグメントが未登録です。現場表を取り込んでください")
    return site_dict, site_master, warnings


def get_manager_contract_codes(manager_id):
    """現場リストPLUSの最新マスタから担当者の契約コード一覧を返す。"""
    ensure_contract_master_synced()
    rows = (
        SiteContractMaster.query
        .filter(SiteContractMaster.is_active.is_(True))
        .order_by(SiteContractMaster.contract_code.asc())
        .all()
    )

    contract_codes = set()
    for row in rows:
        if not manager_ids_match(row.site_manager_id, manager_id):
            continue
        contract_code = normalize_contract_code(row.contract_code)
        if not contract_code:
            continue
        contract_codes.add(contract_code)
    return contract_codes


@subject_analysis_tool_bp.route("/")
@login_required
def index():
    return render_template("subject_analysis_tool.html")


@subject_analysis_tool_bp.route("/api/upload", methods=["POST"])
@login_required
def upload_files():
    try:
        subject_file = request.files.get("subject_file")
        prev_year_subject_file = request.files.get("prev_year_subject_file")
        site_file = request.files.get("site_file")
        site_source = str(request.form.get("site_source", "file") or "file").strip().lower()
        manager_id = str(request.form.get("manager_id", "") or "").strip()

        if not subject_file:
            return jsonify({"error": "科目別分析表 CSV を選択してください"}), 400
        if site_source not in {"file", "db"}:
            return jsonify({"error": "site_source は file または db を指定してください"}), 400
        if site_source == "file" and (not site_file or not site_file.filename):
            return jsonify({"error": "現場表読み込みモードでは現行現場表 CSV が必要です"}), 400

        # 担当者絞り込みは現場リストPLUSの最新マスタを参照する（どちらのモードでも共通）
        manager_contract_codes = None
        if manager_id:
            manager_contract_codes = get_manager_contract_codes(manager_id)
            if not manager_contract_codes:
                return jsonify({
                    "error": f"担当者番号 {manager_id} に紐づく現場が現場リストPLUSに見つかりません。"
                             "担当者番号を確認するか、現場リストPLUSを同期してください"
                }), 400

        upload_folder = get_upload_folder()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        user_id = current_user.username

        subject_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_subject.csv")
        subject_file.save(subject_path)

        subject_data = read_csv_with_encoding(subject_path)
        if not subject_data:
            return jsonify({"error": "科目別分析表 CSV の読み込みに失敗しました"}), 400

        structure_validation = validate_subject_csv_structure(subject_data, "科目別分析表")
        if structure_validation["fatal"]:
            try:
                os.remove(subject_path)
            except OSError:
                pass
            return jsonify({"error": structure_validation["message"]}), 400

        parsed_data = parse_subject_data(subject_data)

        extra_warnings = []
        manager_filter = None
        if manager_contract_codes is not None:
            total_rows = len(parsed_data)
            parsed_data = [item for item in parsed_data if item["contract_code"] in manager_contract_codes]
            if not parsed_data:
                try:
                    os.remove(subject_path)
                except OSError:
                    pass
                return jsonify({
                    "error": f"担当者番号 {manager_id} の現場（{len(manager_contract_codes)}件）は"
                             "科目別分析表に含まれていません"
                }), 400
            manager_filter = {
                "manager_id": manager_id,
                "master_contract_count": len(manager_contract_codes),
                "matched_contract_count": len({item["contract_code"] for item in parsed_data}),
                "excluded_row_count": total_rows - len(parsed_data),
            }
            extra_warnings.append(
                f"担当者番号 {manager_id} で絞り込み: 対象現場 {manager_filter['matched_contract_count']} 件 / "
                f"対象外の行 {manager_filter['excluded_row_count']} 件を除外しました"
            )

        prev_year_data = None
        prev_year_validation = None
        if prev_year_subject_file and prev_year_subject_file.filename:
            prev_year_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_prev_subject.csv")
            prev_year_subject_file.save(prev_year_path)
            prev_year_csv = read_csv_with_encoding(prev_year_path)
            if not prev_year_csv:
                extra_warnings.append("前年データ CSV の読み込みに失敗したため、前年比較なしで続行します")
            if prev_year_csv:
                prev_year_validation = validate_subject_csv_structure(prev_year_csv, "前年データ")
                if prev_year_validation["fatal"]:
                    try:
                        os.remove(prev_year_path)
                    except OSError:
                        pass
                    try:
                        os.remove(subject_path)
                    except OSError:
                        pass
                    return jsonify({"error": prev_year_validation["message"]}), 400
                prev_year_data = parse_subject_data(prev_year_csv)
                if prev_year_data and manager_contract_codes is not None:
                    prev_year_data = [
                        item for item in prev_year_data
                        if item["contract_code"] in manager_contract_codes
                    ]
            try:
                os.remove(prev_year_path)
            except OSError:
                pass

        warnings = []
        if site_source == "db":
            site_data, site_master, warnings = load_site_mapping_from_master()
        else:
            site_data = {}
            site_master = {}
            site_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_site.csv")
            site_file.save(site_path)
            site_csv = read_csv_with_encoding(site_path)
            if site_csv:
                site_data = parse_site_data(site_csv)
            else:
                extra_warnings.append("現行現場表 CSV の読み込みに失敗したため、セグメントは未分類になります")
            try:
                os.remove(site_path)
            except OSError:
                pass

        validation = build_upload_validation(
            subject_data,
            parsed_data,
            site_data,
            prev_year_data or [],
            structure_validation,
            prev_year_validation,
            extra_warnings + list(warnings or []),
        )

        try:
            os.remove(subject_path)
        except OSError:
            pass

        return jsonify(
            {
                "success": True,
                "data": {
                    "current_year": parsed_data,
                    "prev_year": prev_year_data,
                    "site_mapping": site_data,
                    "site_master": site_master,
                },
                "metadata": {
                    "current_year_count": len(parsed_data),
                    "prev_year_count": len(prev_year_data) if prev_year_data else 0,
                    "site_count": len(site_data) if site_data else 0,
                    "site_source": site_source,
                    "manager_filter": manager_filter,
                    "unclassified_count": validation["unmatched_site_mapping_count"],
                    "warnings": validation["warnings"],
                    "validation": validation,
                },
            }
        )

    except Exception as exc:
        traceback.print_exc()
        return jsonify({"error": f"処理中にエラーが発生しました: {exc}"}), 500


@subject_analysis_tool_bp.route("/api/manager_contracts", methods=["GET"])
@login_required
def manager_contracts():
    try:
        manager_id = str(request.args.get("manager_id", "") or "").strip()
        if not manager_id:
            return jsonify({"error": "manager_id を指定してください"}), 400

        contract_codes = get_manager_contract_codes(manager_id)

        return jsonify(
            {
                "manager_id": manager_id,
                "contract_codes": sorted(contract_codes),
                "count": len(contract_codes),
            }
        )
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"error": f"処理中にエラーが発生しました: {exc}"}), 500


def read_csv_with_encoding(file_path):
    """複数エンコーディングで CSV を読み込む。"""
    # utf-8 より先に utf-8-sig を試さないと BOM が先頭セルに残る
    encodings = ["utf-8-sig", "cp932", "shift_jis", "utf-8"]

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding, newline="") as file:
                reader = csv.reader(file)
                return list(reader)
        except Exception:
            continue

    return None


def validate_subject_csv_structure(csv_data, label):
    """科目別分析CSVの最低限の列構造を検証する。"""
    if not csv_data:
        return {"fatal": True, "message": f"{label} CSV が空です"}

    header = csv_data[0] if csv_data else []
    if len(header) < 13:
        return {
            "fatal": True,
            "message": f"{label} CSV の必須列が不足しています。少なくとも13列必要です。",
        }

    short_row_count = 0
    for row in csv_data[1:]:
        if row and len(row) < 13:
            short_row_count += 1

    return {
        "fatal": False,
        "message": "",
        "short_row_count": short_row_count,
        "total_rows": max(len(csv_data) - 1, 0),
    }


def build_upload_validation(
    subject_csv,
    parsed_data,
    site_mapping,
    prev_year_data,
    structure_validation,
    prev_year_validation,
    site_warnings,
):
    """CSV読込結果の検証情報を画面表示用metadataとして組み立てる。"""
    contract_missing_count = 0
    amount_conversion_error_count = 0

    for index, row in enumerate(subject_csv):
        if index == 0:
            continue
        if len(row) < 13:
            continue
        if not normalize_contract_code(row[8] if len(row) > 8 else ""):
            contract_missing_count += 1
        for value in row[13:]:
            text = str(value or "").strip().replace(",", "")
            if not text:
                continue
            try:
                float(text)
            except ValueError:
                amount_conversion_error_count += 1

    parsed_codes = {item["contract_code"] for item in parsed_data if item.get("contract_code")}
    unmatched_site_mapping_count = len([code for code in parsed_codes if code not in site_mapping])

    prev_year_keys = {
        (item.get("contract_code"), item.get("subject_name"))
        for item in prev_year_data
        if item.get("contract_code") and item.get("subject_name")
    }
    current_keys = {
        (item.get("contract_code"), item.get("subject_name"))
        for item in parsed_data
        if item.get("contract_code") and item.get("subject_name")
    }
    prev_year_missing_count = 0
    if prev_year_data:
        prev_year_missing_count = len([key for key in current_keys if key not in prev_year_keys])

    warnings = []
    if structure_validation.get("short_row_count"):
        warnings.append(f"列不足のためスキップされた行が {structure_validation['short_row_count']} 件あります")
    if contract_missing_count:
        warnings.append(f"契約コードが空の行が {contract_missing_count} 件あります")
    if amount_conversion_error_count:
        warnings.append(f"金額へ変換できないセルが {amount_conversion_error_count} 件あります。該当セルは0として扱われます")
    if unmatched_site_mapping_count:
        warnings.append(f"現場マッピングに一致しない契約コードが {unmatched_site_mapping_count} 件あります")
    if prev_year_missing_count:
        warnings.append(f"前年データに対応行がない比較対象が {prev_year_missing_count} 件あります")
    if prev_year_validation and prev_year_validation.get("short_row_count"):
        warnings.append(f"前年データで列不足の行が {prev_year_validation['short_row_count']} 件あります")
    warnings.extend(site_warnings or [])

    return {
        "row_count": len(parsed_data),
        "prev_year_row_count": len(prev_year_data),
        "site_mapping_count": len(site_mapping or {}),
        "unmatched_site_mapping_count": unmatched_site_mapping_count,
        "contract_missing_count": contract_missing_count,
        "amount_conversion_error_count": amount_conversion_error_count,
        "prev_year_missing_comparison_count": prev_year_missing_count,
        "short_row_count": structure_validation.get("short_row_count", 0),
        "warnings": warnings,
    }


def parse_amount_cell(value):
    text = str(value or "").strip().replace(",", "")
    if not text:
        return 0
    return float(text)


def parse_subject_data(csv_data):
    """科目別分析表 CSV を解析して明細配列に変換する。

    符号の扱い: CSV 上は売上・原価とも正の値。原価行のみ符号反転して
    負の値で保持するため、利益 = 売上 + 原価（足し算）で計算できる。
    """
    parsed = []
    indirect_cost_codes = {"間接原価", "関節原価"}
    revenue_names = {"基本請負料", "その他請負料"}
    # 請負形態名称で売上区分を判別する科目（月次ジェネレーターと同じ扱い）
    contract_type_sales_subjects = {"自動車売上", "旅客運送売上"}

    for index, row in enumerate(csv_data):
        if index == 0:
            continue
        if len(row) < 13:
            continue

        contract_code = normalize_contract_code(row[8] if len(row) > 8 else "")
        corp_name = row[9].strip() if len(row) > 9 else ""
        site_name = row[10].strip() if len(row) > 10 else ""
        subject_code = row[11].strip() if len(row) > 11 else ""
        subject_name = row[12].strip() if len(row) > 12 else ""
        contract_type = row[5].strip() if len(row) > 5 else ""

        normalized_subject_code = subject_code.replace(" ", "").replace("　", "")
        is_indirect_cost = normalized_subject_code in indirect_cost_codes

        # 販売費は現場収支の対象外（月次ジェネレーターと同じ扱い）
        if normalized_subject_code == "販売費":
            continue
        if not subject_name and not is_indirect_cost:
            continue
        if not contract_code:
            continue

        display_subject_name = subject_name
        if not display_subject_name and is_indirect_cost:
            display_subject_name = "間接原価"

        is_contract_type_sales = subject_name in contract_type_sales_subjects
        if is_contract_type_sales:
            if contract_type in revenue_names:
                display_subject_name = contract_type
            elif contract_type:
                display_subject_name = f"{subject_name}({contract_type})"

        amounts = []
        for col_index in range(13, len(row)):
            try:
                amounts.append(parse_amount_cell(row[col_index]))
            except Exception:
                amounts.append(0)

        while len(amounts) < 12:
            amounts.append(0)

        is_revenue = display_subject_name in revenue_names or is_contract_type_sales
        if not is_revenue:
            amounts = [-amount for amount in amounts]

        parsed.append(
            {
                "contract_code": contract_code,
                "corp_name": corp_name,
                "site_name": site_name,
                "subject_code": subject_code,
                "subject_name": display_subject_name,
                "original_subject_name": subject_name,
                "contract_type": contract_type,
                "is_revenue": is_revenue,
                "amounts": amounts[:12],
            }
        )

    return parsed


def parse_site_data(csv_data):
    """現場表 CSV を解析して 契約コード -> セグメント に変換する。"""
    site_dict = {}

    for index, row in enumerate(csv_data):
        if index == 0:
            continue
        if len(row) < 2:
            continue

        contract_code = normalize_contract_code(row[0])
        segment = row[1].strip()
        if not contract_code:
            continue

        site_dict[contract_code] = segment

    return site_dict


def summarize_by_key(results, key_getter, label_getter=None):
    summary = {}
    for result in results:
        key = key_getter(result)
        if key not in summary:
            summary[key] = {
                "name": label_getter(result) if label_getter else key,
                "current": 0,
                "comparison": 0,
                "diff": 0,
                "anomaly_count": 0,
            }
        summary[key]["current"] += float(result.get("currentValue", 0) or 0)
        summary[key]["comparison"] += float(result.get("comparisonValue", 0) or 0)
        summary[key]["diff"] += float(result.get("diff", 0) or 0)
        if result.get("isAnomaly"):
            summary[key]["anomaly_count"] += 1

    for row in summary.values():
        row["diff_rate"] = (row["diff"] / row["comparison"] * 100) if row["comparison"] else 0
    return sorted(summary.values(), key=lambda item: abs(item["diff"]), reverse=True)


def append_sheet(sheet, headers, rows, number_columns=None, percent_columns=None):
    number_columns = set(number_columns or [])
    percent_columns = set(percent_columns or [])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2B4058")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row, start=1):
            if index in number_columns:
                cell.number_format = '#,##0'
            if index in percent_columns:
                cell.number_format = '0.00'

    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


@subject_analysis_tool_bp.route("/api/export/xlsx", methods=["POST"])
@login_required
def export_xlsx():
    try:
        payload = request.get_json(silent=True) or {}
        results = payload.get("results") or []
        filters = payload.get("filters") or {}
        if not results:
            return jsonify({"error": "エクスポートするデータがありません"}), 400

        workbook = Workbook()
        detail_sheet = workbook.active
        detail_sheet.title = "詳細一覧"

        detail_rows = []
        for result in results:
            item = result.get("item") or {}
            detail_rows.append([
                item.get("contract_code", ""),
                item.get("site_name", ""),
                item.get("corp_name", ""),
                result.get("segment", "未分類"),
                item.get("subject_name", ""),
                f"{result.get('month')}月",
                result.get("currentValue", 0),
                result.get("comparisonValue", 0) if result.get("hasComparison") else None,
                result.get("diff", 0) if result.get("hasComparison") else None,
                result.get("diffRate", 0) if result.get("hasComparison") else None,
            ])

        append_sheet(
            detail_sheet,
            ["契約コード", "現場名", "法人名", "セグメント", "科目名", "月", "当期値", "比較値", "差異", "差異率(%)"],
            detail_rows,
            number_columns={7, 8, 9},
            percent_columns={10},
        )

        site_sheet = workbook.create_sheet("現場別サマリー")
        # 同名現場の混同を避けるため契約コードを含めたキーで集計する
        site_summary = summarize_by_key(
            results,
            lambda r: (
                (r.get("item") or {}).get("contract_code", ""),
                (r.get("item") or {}).get("site_name", ""),
            ),
            lambda r: (r.get("item") or {}).get("site_name", ""),
        )
        append_sheet(
            site_sheet,
            ["現場名", "当期合計", "比較合計", "差異", "差異率(%)", "異常値件数"],
            [[r["name"], r["current"], r["comparison"], r["diff"], r["diff_rate"], r["anomaly_count"]] for r in site_summary],
            number_columns={2, 3, 4, 6},
            percent_columns={5},
        )

        subject_sheet = workbook.create_sheet("科目別サマリー")
        subject_summary = summarize_by_key(results, lambda r: (r.get("item") or {}).get("subject_name", ""))
        append_sheet(
            subject_sheet,
            ["科目名", "当期合計", "比較合計", "差異", "差異率(%)", "異常値件数"],
            [[r["name"], r["current"], r["comparison"], r["diff"], r["diff_rate"], r["anomaly_count"]] for r in subject_summary],
            number_columns={2, 3, 4, 6},
            percent_columns={5},
        )

        condition_sheet = workbook.create_sheet("条件")
        condition_rows = [[key, ", ".join(map(str, value)) if isinstance(value, list) else value] for key, value in filters.items()]
        append_sheet(condition_sheet, ["項目", "値"], condition_rows)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"科目別分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"error": f"Excel出力中にエラーが発生しました: {exc}"}), 500
