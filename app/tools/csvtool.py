from __future__ import annotations

import csv
import io
import re
import sqlite3
import tempfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import chardet
from flask import Blueprint, jsonify, render_template, request, send_file
from flask_login import login_required

csvtool_bp = Blueprint("csvtool", __name__, url_prefix="/tools/csvtool")

# Safety caps for table / database / spreadsheet ingestion.
MAX_ROWS = 200_000
MAX_COLS = 1_024


def column_letter(index: int) -> str:
    """0-based column index -> Excel style letters (0->A, 25->Z, 26->AA)."""
    if index < 0:
        return ""
    letters = ""
    index += 1
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

ALLOWED_DELIMITERS = {
    ",": "comma",
    "\t": "tab",
    ";": "semicolon",
    "|": "pipe",
}
ALLOWED_ENCODINGS = {
    "utf-8",
    "utf-8-sig",
    "shift_jis",
    "cp932",
    "euc-jp",
    "iso-2022-jp",
    "latin-1",
}
FORMULA_PREFIXES = ("=", "+", "-", "@")
DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y")


@csvtool_bp.route("/")
@login_required
def csvtool():
    return render_template("csvtool.html")


def _decode_csv_bytes(file_bytes: bytes, requested_encoding: str | None = None) -> tuple[str, str, float]:
    detected = chardet.detect(file_bytes)
    detected_encoding = detected.get("encoding") or "utf-8"
    confidence = float(detected.get("confidence") or 0)

    candidates = []
    if requested_encoding:
        candidates.append(requested_encoding)
    candidates.extend([detected_encoding, "utf-8-sig", "utf-8", "cp932", "shift_jis", "euc-jp", "iso-2022-jp", "latin-1"])

    seen: set[str] = set()
    last_error: Exception | None = None
    for encoding in candidates:
        key = (encoding or "").lower()
        if not key or key in seen:
            continue
        seen.add(key)
        try:
            text = file_bytes.decode(encoding)
            if text.startswith("\ufeff"):
                text = text[1:]
            return text, encoding, confidence
        except (UnicodeDecodeError, LookupError) as exc:
            last_error = exc

    raise ValueError(f"エンコーディングの検出に失敗しました: {last_error}")


def _safe_sniff_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:20])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="".join(ALLOWED_DELIMITERS))
        if dialect.delimiter in ALLOWED_DELIMITERS:
            return dialect.delimiter
    except csv.Error:
        pass

    counts = Counter()
    for line in sample.splitlines():
        if not line.strip():
            continue
        for delimiter in ALLOWED_DELIMITERS:
            counts[delimiter] += line.count(delimiter)
    return counts.most_common(1)[0][0] if counts else ","


def _parse_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _normalize_delimiter(raw: str | None, text: str) -> str:
    if not raw or raw == "auto":
        return _safe_sniff_delimiter(text)
    if raw == "\\t":
        raw = "\t"
    if raw not in ALLOWED_DELIMITERS:
        raise ValueError("区切り文字が不正です。")
    return raw


def _build_headers(rows: list[list[str]], has_header: bool, width: int) -> tuple[list[str], list[list[str]]]:
    if has_header and rows:
        headers = rows[0][:]
        data = rows[1:]
    else:
        headers = []
        data = rows[:]

    if len(headers) < width:
        headers.extend([column_letter(idx) for idx in range(len(headers), width)])
    elif len(headers) > width:
        width = len(headers)

    # Replace any blank header cell with its Excel-style column letter.
    headers = [(h if str(h).strip() != "" else column_letter(idx)) for idx, h in enumerate(headers[:width])]
    return headers, data


def _normalize_rows(
    rows: list[list[str]],
    width: int,
    *,
    preserve_empty_rows: bool,
    row_width_strategy: str,
) -> tuple[list[list[str]], list[dict]]:
    data: list[list[str]] = []
    warnings: list[dict] = []
    for index, row in enumerate(rows):
        if not preserve_empty_rows and not any(str(cell).strip() for cell in row):
            continue
        original_width = len(row)
        normalized = row[:]
        if original_width < width:
            normalized.extend([""] * (width - original_width))
            if original_width != width:
                warnings.append({"row": index, "type": "padded", "message": f"{index + 1}行目を{width}列に補完しました。"})
        elif original_width > width:
            if row_width_strategy == "error":
                warnings.append({"row": index, "type": "width", "message": f"{index + 1}行目の列数が多すぎます。"})
            elif row_width_strategy == "trim":
                normalized = normalized[:width]
                warnings.append({"row": index, "type": "trimmed", "message": f"{index + 1}行目を{width}列に切り詰めました。"})
            # "expand" / "pad": keep the wider row as-is (width was already
            # widened to the max in _parse_csv_upload for "expand").
        data.append(normalized)
    return data, warnings


def _parse_csv_upload(file_bytes: bytes, options: dict) -> dict:
    text, used_encoding, confidence = _decode_csv_bytes(file_bytes, options.get("encoding"))
    delimiter = _normalize_delimiter(options.get("delimiter"), text)

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise ValueError("CSVにデータがありません。")

    has_header = _parse_bool(options.get("has_header"), True)
    preserve_empty_rows = _parse_bool(options.get("preserve_empty_rows"), True)
    row_width_strategy = options.get("row_width_strategy") or "expand"
    if row_width_strategy not in {"expand", "pad", "trim", "error"}:
        raise ValueError("行幅の扱いが不正です。")

    max_width = max((len(row) for row in rows), default=0)
    if has_header:
        width = len(rows[0])
        if row_width_strategy == "expand":
            width = max_width
    else:
        width = max_width

    headers, raw_data = _build_headers(rows, has_header, width)
    data, warnings = _normalize_rows(
        raw_data,
        len(headers),
        preserve_empty_rows=preserve_empty_rows,
        row_width_strategy=row_width_strategy,
    )

    return {
        "headers": headers,
        "data": data,
        "encoding": used_encoding,
        "confidence": round(confidence * 100, 1),
        "delimiter": delimiter,
        "hasHeader": has_header,
        "preserveEmptyRows": preserve_empty_rows,
        "rowWidthStrategy": row_width_strategy,
        "totalRows": len(data),
        "totalColumns": len(headers),
        "warnings": warnings[:200],
        "previewRows": rows[:20],
        "rawRowCount": len(rows),
        "maxColumns": max_width,
        "source": "csv",
    }


@csvtool_bp.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "ファイルが選択されていません。"}), 400

    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "ファイルが空です。"}), 400

    try:
        parsed = _parse_csv_upload(file_bytes, request.form.to_dict())
    except (ValueError, csv.Error) as exc:
        return jsonify({"error": str(exc)}), 400

    parsed["filename"] = file.filename
    return jsonify(parsed)


# ===================== SQLite database support =====================

def _open_sqlite_readonly(file_bytes: bytes) -> tuple[sqlite3.Connection, str]:
    """Persist uploaded bytes to a temp file and open it strictly read-only."""
    if not file_bytes.startswith(b"SQLite format 3\x00"):
        raise ValueError("SQLite データベースファイルではありません。")
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    try:
        tmp.write(file_bytes)
        tmp.flush()
        tmp.close()
        uri = f"file:{Path(tmp.name).as_posix()}?mode=ro&immutable=1"
        conn = sqlite3.connect(uri, uri=True)
        conn.execute("PRAGMA query_only = ON")
        return conn, tmp.name
    except Exception:
        Path(tmp.name).unlink(missing_ok=True)
        raise


def _quote_ident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _list_sqlite_tables(conn: sqlite3.Connection) -> list[dict]:
    cur = conn.execute(
        "SELECT name, type FROM sqlite_master "
        "WHERE type IN ('table','view') AND name NOT LIKE 'sqlite_%' ORDER BY type, name"
    )
    tables = []
    for name, obj_type in cur.fetchall():
        try:
            count = conn.execute(f"SELECT COUNT(*) FROM {_quote_ident(name)}").fetchone()[0]
        except sqlite3.Error:
            count = None
        try:
            cols = [row[1] for row in conn.execute(f"PRAGMA table_info({_quote_ident(name)})").fetchall()]
        except sqlite3.Error:
            cols = []
        tables.append({"name": name, "type": obj_type, "rows": count, "columns": cols})
    return tables


@csvtool_bp.route("/api/db/tables", methods=["POST"])
@login_required
def api_db_tables():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "ファイルが選択されていません。"}), 400
    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "ファイルが空です。"}), 400

    conn = None
    tmp_path = None
    try:
        conn, tmp_path = _open_sqlite_readonly(file_bytes)
        tables = _list_sqlite_tables(conn)
    except (ValueError, sqlite3.Error) as exc:
        return jsonify({"error": f"データベースを開けませんでした: {exc}"}), 400
    finally:
        if conn is not None:
            conn.close()
        if tmp_path:
            Path(tmp_path).unlink(missing_ok=True)

    if not tables:
        return jsonify({"error": "読み込めるテーブルがありません。"}), 400
    return jsonify({"tables": tables, "filename": file.filename})


@csvtool_bp.route("/api/db/load", methods=["POST"])
@login_required
def api_db_load():
    file = request.files.get("file")
    table = request.form.get("table", "")
    if not file:
        return jsonify({"error": "ファイルが選択されていません。"}), 400
    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "ファイルが空です。"}), 400

    conn = None
    tmp_path = None
    truncated = False
    try:
        conn, tmp_path = _open_sqlite_readonly(file_bytes)
        valid_names = {t["name"] for t in _list_sqlite_tables(conn)}
        if table not in valid_names:
            return jsonify({"error": "指定されたテーブルが見つかりません。"}), 400

        cur = conn.execute(f"SELECT * FROM {_quote_ident(table)} LIMIT {MAX_ROWS + 1}")
        headers = [d[0] for d in cur.description] if cur.description else []
        if len(headers) > MAX_COLS:
            return jsonify({"error": f"列数が多すぎます (上限 {MAX_COLS})。"}), 400
        rows = cur.fetchall()
        truncated = len(rows) > MAX_ROWS
        rows = rows[:MAX_ROWS]
        data = [["" if v is None else str(v) for v in row] for row in rows]
    except (ValueError, sqlite3.Error) as exc:
        return jsonify({"error": f"テーブルを読み込めませんでした: {exc}"}), 400
    finally:
        if conn is not None:
            conn.close()
        if tmp_path:
            Path(tmp_path).unlink(missing_ok=True)

    headers = [(h if str(h).strip() != "" else column_letter(i)) for i, h in enumerate(headers)]
    return jsonify({
        "headers": headers,
        "data": data,
        "encoding": "utf-8-sig",
        "delimiter": ",",
        "filename": f"{Path(file.filename or 'database').stem}__{table}.csv",
        "tableName": table,
        "totalRows": len(data),
        "totalColumns": len(headers),
        "truncated": truncated,
        "source": "sqlite",
    })


# ===================== Excel (.xlsx) support =====================

def _load_workbook(file_bytes: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency declared in requirements
        raise ValueError("Excel 読み込みライブラリ (openpyxl) が利用できません。") from exc
    return load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)


def _xlsx_cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d %H:%M:%S")
        return text[:-9] if text.endswith(" 00:00:00") else text
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


@csvtool_bp.route("/api/xlsx/sheets", methods=["POST"])
@login_required
def api_xlsx_sheets():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "ファイルが選択されていません。"}), 400
    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "ファイルが空です。"}), 400
    try:
        wb = _load_workbook(file_bytes)
        sheets = [{"name": ws.title, "rows": ws.max_row or 0, "columns": ws.max_column or 0}
                  for ws in wb.worksheets]
        wb.close()
    except Exception as exc:
        return jsonify({"error": f"Excelファイルを開けませんでした: {exc}"}), 400
    if not sheets:
        return jsonify({"error": "シートがありません。"}), 400
    return jsonify({"sheets": sheets, "filename": file.filename})


@csvtool_bp.route("/api/xlsx/load", methods=["POST"])
@login_required
def api_xlsx_load():
    file = request.files.get("file")
    sheet = request.form.get("sheet", "")
    has_header = _parse_bool(request.form.get("has_header"), True)
    if not file:
        return jsonify({"error": "ファイルが選択されていません。"}), 400
    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "ファイルが空です。"}), 400

    try:
        wb = _load_workbook(file_bytes)
        if sheet and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]
            sheet = ws.title

        rows: list[list[str]] = []
        for excel_row in ws.iter_rows(values_only=True):
            rows.append([_xlsx_cell_to_str(v) for v in excel_row])
            if len(rows) > MAX_ROWS:
                break
        wb.close()
    except Exception as exc:
        return jsonify({"error": f"シートを読み込めませんでした: {exc}"}), 400

    while rows and not any(str(c).strip() for c in rows[-1]):
        rows.pop()
    if not rows:
        return jsonify({"error": "シートにデータがありません。"}), 400

    width = max(len(r) for r in rows)
    if width > MAX_COLS:
        return jsonify({"error": f"列数が多すぎます (上限 {MAX_COLS})。"}), 400
    headers, raw_data = _build_headers(rows, has_header, width)
    data, warnings = _normalize_rows(raw_data, len(headers), preserve_empty_rows=True, row_width_strategy="pad")

    return jsonify({
        "headers": headers,
        "data": data,
        "encoding": "utf-8-sig",
        "delimiter": ",",
        "filename": f"{Path(file.filename or 'book').stem}__{sheet}.csv",
        "sheetName": sheet,
        "totalRows": len(data),
        "totalColumns": len(headers),
        "warnings": warnings[:200],
        "source": "xlsx",
    })


def _sanitize_for_csv(value: object, *, protect_formulas: bool) -> str:
    text = "" if value is None else str(value)
    if protect_formulas and text.startswith(FORMULA_PREFIXES):
        return "'" + text
    return text


@csvtool_bp.route("/api/download", methods=["POST"])
@login_required
def api_download():
    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({"error": "データがありません。"}), 400

    headers = payload.get("headers", [])
    data = payload.get("data", [])
    encoding = str(payload.get("encoding", "utf-8-sig")).lower()
    delimiter = payload.get("delimiter", ",")
    filename = payload.get("filename", "power_csv_export.csv")
    protect_formulas = _parse_bool(payload.get("protectFormulas"), True)

    if delimiter == "\\t":
        delimiter = "\t"
    if delimiter not in ALLOWED_DELIMITERS:
        return jsonify({"error": "区切り文字が不正です。"}), 400
    if encoding not in ALLOWED_ENCODINGS:
        return jsonify({"error": "エンコーディングが不正です。"}), 400

    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, lineterminator="\r\n")
    writer.writerow([_sanitize_for_csv(cell, protect_formulas=protect_formulas) for cell in headers])
    for row in data:
        writer.writerow([_sanitize_for_csv(cell, protect_formulas=protect_formulas) for cell in row])

    try:
        raw = output.getvalue().encode(encoding)
    except LookupError:
        return jsonify({"error": "エンコーディングが不正です。"}), 400
    except UnicodeEncodeError as exc:
        return jsonify({"error": f"指定エンコーディングで保存できない文字があります: {exc.object[exc.start:exc.end]}"}), 400

    return send_file(
        io.BytesIO(raw),
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv; charset=utf-8",
    )


@csvtool_bp.route("/api/export/xlsx", methods=["POST"])
@login_required
def api_export_xlsx():
    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({"error": "データがありません。"}), 400
    headers = payload.get("headers", [])
    data = payload.get("data", [])
    filename = payload.get("filename", "power_csv_export.xlsx")
    sheet_name = str(payload.get("sheetName") or "Sheet1")[:31] or "Sheet1"

    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"error": "Excel 書き出しライブラリ (openpyxl) が利用できません。"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([str(h) for h in headers])
    for row in data:
        ws.append([("" if c is None else str(c)) for c in row])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _sanitize_sql_identifier(name: str, fallback: str) -> str:
    # Identifiers are always double-quoted, so Unicode (incl. Japanese) is fine.
    # Strip only control characters / quotes and collapse whitespace.
    cleaned = re.sub(r'["\x00-\x1f\x7f]', "", str(name or "")).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    if not cleaned:
        return fallback
    if cleaned[0].isdigit():
        cleaned = f"col_{cleaned}"
    return cleaned


@csvtool_bp.route("/api/export/db", methods=["POST"])
@login_required
def api_export_db():
    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({"error": "データがありません。"}), 400
    headers = payload.get("headers", [])
    data = payload.get("data", [])
    filename = payload.get("filename", "power_csv_export.db")
    table_name = _sanitize_sql_identifier(payload.get("tableName") or "data", "data")

    if not headers:
        return jsonify({"error": "ヘッダーがありません。"}), 400

    # Build unique, SQL-safe column names.
    col_names: list[str] = []
    used: set[str] = set()
    for idx, h in enumerate(headers):
        base = _sanitize_sql_identifier(h, column_letter(idx))
        candidate = base
        suffix = 1
        while candidate.lower() in used:
            suffix += 1
            candidate = f"{base}_{suffix}"
        used.add(candidate.lower())
        col_names.append(candidate)

    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    try:
        conn = sqlite3.connect(tmp.name)
        cols_ddl = ", ".join(f"{_quote_ident(c)} TEXT" for c in col_names)
        conn.execute(f"CREATE TABLE {_quote_ident(table_name)} ({cols_ddl})")
        placeholders = ", ".join(["?"] * len(col_names))
        insert_sql = f"INSERT INTO {_quote_ident(table_name)} VALUES ({placeholders})"
        width = len(col_names)
        for row in data:
            row = list(row)[:width] + [""] * (width - len(row))
            conn.execute(insert_sql, ["" if c is None else str(c) for c in row])
        conn.commit()
        conn.close()
        raw = Path(tmp.name).read_bytes()
    finally:
        Path(tmp.name).unlink(missing_ok=True)

    if not filename.lower().endswith((".db", ".sqlite", ".sqlite3")):
        filename += ".db"
    return send_file(
        io.BytesIO(raw),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.sqlite3",
    )


def _is_number(value: str) -> bool:
    if str(value).strip() == "":
        return False
    try:
        float(str(value).replace(",", ""))
        return True
    except ValueError:
        return False


def _is_date(value: str) -> bool:
    text = str(value).strip()
    if not text:
        return False
    for fmt in DATE_FORMATS:
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def _add_issue(issues: list[dict], *, issue_type: str, severity: str, row: int, col: int, message: str) -> None:
    issues.append({
        "type": issue_type,
        "severity": severity,
        "row": row,
        "col": col,
        "message": message,
    })


def _validate_custom_rule(rule: dict, headers: list[str], data: list[list[str]], issues: list[dict]) -> None:
    try:
        col = int(rule.get("col"))
    except (TypeError, ValueError):
        return
    if col < 0 or col >= len(headers):
        return

    rule_type = rule.get("type")
    severity = rule.get("severity", "error")
    value = str(rule.get("value", ""))
    header = headers[col] or column_letter(col)
    seen: dict[str, int] = {}
    pattern = None
    if rule_type == "regex":
        try:
            pattern = re.compile(value)
        except re.error:
            _add_issue(issues, issue_type="rule", severity="error", row=-1, col=col, message=f"{header}: 正規表現が不正です。")
            return

    choices = [item.strip() for item in value.split(",") if item.strip()]
    for row_index, row in enumerate(data):
        cell = str(row[col]) if col < len(row) else ""
        stripped = cell.strip()
        if rule_type == "required" and not stripped:
            _add_issue(issues, issue_type="required", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」は必須です。")
        elif rule_type == "unique" and stripped:
            if stripped in seen:
                _add_issue(issues, issue_type="unique", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」は{seen[stripped] + 2}行目と重複しています。")
            else:
                seen[stripped] = row_index
        elif rule_type == "number" and stripped and not _is_number(stripped):
            _add_issue(issues, issue_type="number", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」は数値ではありません。")
        elif rule_type == "integer" and stripped and not re.fullmatch(r"[+-]?\d+", stripped.replace(",", "")):
            _add_issue(issues, issue_type="integer", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」は整数ではありません。")
        elif rule_type == "date" and stripped and not _is_date(stripped):
            _add_issue(issues, issue_type="date", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」は日付ではありません。")
        elif rule_type == "regex" and stripped and pattern and not pattern.search(stripped):
            _add_issue(issues, issue_type="regex", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」はパターンに一致しません。")
        elif rule_type == "length" and stripped:
            try:
                expected = int(value)
            except ValueError:
                continue
            if len(stripped) != expected:
                _add_issue(issues, issue_type="length", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」は{expected}文字ではありません。")
        elif rule_type == "enum" and stripped and choices and stripped not in choices:
            _add_issue(issues, issue_type="enum", severity=severity, row=row_index, col=col, message=f"{row_index + 2}行目「{header}」は候補値にありません。")


@csvtool_bp.route("/api/validate", methods=["POST"])
@login_required
def api_validate():
    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({"error": "データがありません。"}), 400

    headers = payload.get("headers", [])
    data = payload.get("data", [])
    rules = payload.get("rules", [])
    column_types = payload.get("columnTypes", {})
    issues: list[dict] = []

    col_stats = {}
    for col_idx, header in enumerate(headers):
        col_values = [str(row[col_idx]) if col_idx < len(row) else "" for row in data]
        non_empty = [v for v in col_values if v.strip()]
        unique_values = set(non_empty)
        empty_count = len(col_values) - len(non_empty)
        numeric_count = sum(1 for v in non_empty if _is_number(v))
        date_count = sum(1 for v in non_empty if _is_date(v))
        col_stats[header] = {
            "index": col_idx,
            "total": len(col_values),
            "empty": empty_count,
            "unique": len(unique_values),
            "fillRate": round((len(non_empty) / max(len(col_values), 1)) * 100, 1),
            "numericRate": round((numeric_count / max(len(non_empty), 1)) * 100, 1),
            "dateRate": round((date_count / max(len(non_empty), 1)) * 100, 1),
        }

    empty_cells = 0
    for i, row in enumerate(data):
        if len(row) != len(headers):
            _add_issue(issues, issue_type="width", severity="warning", row=i, col=-1, message=f"{i + 2}行目の列数がヘッダーと一致しません。")
        for j, cell in enumerate(row[:len(headers)]):
            if str(cell).strip() == "":
                empty_cells += 1
            locked_type = column_types.get(str(j)) or column_types.get(j)
            if locked_type == "number" and str(cell).strip() and not _is_number(str(cell)):
                _add_issue(issues, issue_type="type", severity="error", row=i, col=j, message=f"{i + 2}行目「{headers[j]}」は数値ロックに違反しています。")
            elif locked_type == "date" and str(cell).strip() and not _is_date(str(cell)):
                _add_issue(issues, issue_type="type", severity="error", row=i, col=j, message=f"{i + 2}行目「{headers[j]}」は日付ロックに違反しています。")
            elif str(cell).startswith(FORMULA_PREFIXES):
                _add_issue(issues, issue_type="formula_injection", severity="warning", row=i, col=j, message=f"{i + 2}行目「{headers[j]}」は表計算ソフトで数式として実行される可能性があります。")

    seen = {}
    duplicate_count = 0
    for i, row in enumerate(data):
        key = tuple(row)
        if key in seen:
            _add_issue(issues, issue_type="duplicate", severity="error", row=i, col=-1, message=f"{i + 2}行目は{seen[key] + 2}行目と重複しています。")
            duplicate_count += 1
        else:
            seen[key] = i

    for rule in rules if isinstance(rules, list) else []:
        if isinstance(rule, dict):
            _validate_custom_rule(rule, headers, data, issues)

    return jsonify({
        "issues": issues,
        "summary": {
            "totalRows": len(data),
            "totalColumns": len(headers),
            "emptyCells": empty_cells,
            "duplicateRows": duplicate_count,
            "issueCount": len(issues),
        },
        "columnStats": col_stats,
    })
