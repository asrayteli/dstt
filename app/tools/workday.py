from flask import Blueprint, render_template, request, send_file
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from flask_login import login_required

try:
    from .japan_holidays import JAPAN_HOLIDAYS
except ImportError:
    from app.tools.japan_holidays import JAPAN_HOLIDAYS  # type: ignore

workday_bp = Blueprint("workday", __name__, url_prefix="/tools/workday")

WEEKDAY_LABELS = ["日", "月", "火", "水", "木", "金", "土"]


def parse_holidays(raw_holidays: str) -> set[str]:
    if not raw_holidays:
        return set()

    normalized = (
        raw_holidays.replace("\n", ",")
        .replace("\r", ",")
        .replace("\t", ",")
        .replace(" ", ",")
        .replace("、", ",")
        .replace("，", ",")
    )

    values = set()
    for token in normalized.split(","):
        token = token.strip()
        if not token:
            continue
        try:
            parsed = datetime.strptime(token, "%Y-%m-%d")
            values.add(parsed.strftime("%Y-%m-%d"))
        except ValueError:
            continue
    return values


def to_sunday_start_weekday(dt: datetime) -> int:
    # datetime.weekday: 月=0..日=6 -> 日=0..土=6
    return (dt.weekday() + 1) % 7


def is_holiday(dt: datetime, holidays: set[str]) -> bool:
    return dt.strftime("%Y-%m-%d") in holidays


def is_workday(dt: datetime, weekdays: set[int], holidays: set[str], count_holidays_as_workdays: bool) -> bool:
    selected_weekday = to_sunday_start_weekday(dt) in weekdays
    holiday = is_holiday(dt, holidays)
    if count_holidays_as_workdays:
        return selected_weekday or holiday
    return selected_weekday and not holiday


def each_day(start_date: datetime, end_date: datetime):
    current = start_date
    while current <= end_date:
        yield current
        current += timedelta(days=1)


def count_workdays(start_date: datetime, end_date: datetime, weekdays: set[int], holidays: set[str], count_holidays_as_workdays: bool) -> int:
    return sum(
        1
        for day in each_day(start_date, end_date)
        if is_workday(day, weekdays, holidays, count_holidays_as_workdays)
    )


def generate_calendar(start_date: datetime, end_date: datetime, weekdays: set[int], holidays: set[str], count_holidays_as_workdays: bool):
    calendar_data = {}
    for day in each_day(start_date, end_date):
        key = (day.year, day.month)
        calendar_data.setdefault(key, []).append(day)

    output = {}
    for (year, month), days_in_month in calendar_data.items():
        first_day = datetime(year, month, 1)
        adjusted_weekday = to_sunday_start_weekday(first_day)
        weeks = []
        week = [None] * adjusted_weekday

        for day in days_in_month:
            week.append(day)
            if len(week) == 7:
                weeks.append(week)
                week = []

        if week:
            week.extend([None] * (7 - len(week)))
            weeks.append(week)

        rendered_weeks = []
        for row in weeks:
            rendered_row = []
            for day in row:
                if day is None:
                    rendered_row.append({"date": None, "is_workday": False, "is_holiday": False})
                    continue
                rendered_row.append(
                    {
                        "date": day,
                        "is_workday": is_workday(day, weekdays, holidays, count_holidays_as_workdays),
                        "is_holiday": is_holiday(day, holidays),
                    }
                )
            rendered_weeks.append(rendered_row)
        output[(year, month)] = rendered_weeks

    return output


def generate_excel_calendar(start_date: datetime, end_date: datetime, weekdays: set[int], holidays: set[str], count_holidays_as_workdays: bool):
    wb = Workbook()
    ws = wb.active
    ws.title = "営業日カレンダー"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    workday_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    holiday_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")

    current_row = 1
    current_date = datetime(start_date.year, start_date.month, 1)
    end_marker = datetime(end_date.year, end_date.month, 1)

    while current_date <= end_marker:
        year = current_date.year
        month = current_date.month

        first_day = datetime(year, month, 1)
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)

        ws.merge_cells(f"A{current_row}:G{current_row}")
        ws[f"A{current_row}"] = f"{year}年{month}月"
        ws[f"A{current_row}"].font = Font(bold=True, size=14)
        ws[f"A{current_row}"].alignment = center
        ws[f"A{current_row}"].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        current_row += 1

        for col, label in enumerate(WEEKDAY_LABELS, 1):
            cell = ws.cell(row=current_row, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center
        current_row += 1

        col = 1
        week_row = current_row
        for _ in range(to_sunday_start_weekday(first_day)):
            cell = ws.cell(row=week_row, column=col, value="")
            cell.border = border
            col += 1

        cursor = first_day
        while cursor <= last_day:
            cell = ws.cell(row=week_row, column=col, value=cursor.day if start_date <= cursor <= end_date else "")
            cell.border = border
            cell.alignment = center
            if start_date <= cursor <= end_date:
                if is_workday(cursor, weekdays, holidays, count_holidays_as_workdays):
                    cell.fill = workday_fill
                else:
                    cell.fill = holiday_fill

            col += 1
            if col > 7:
                col = 1
                week_row += 1
            cursor += timedelta(days=1)

        if col <= 7:
            for rest in range(col, 8):
                cell = ws.cell(row=week_row, column=rest, value="")
                cell.border = border

        current_row = week_row + 2

        if month == 12:
            current_date = datetime(year + 1, 1, 1)
        else:
            current_date = datetime(year, month + 1, 1)

    total_workdays = count_workdays(start_date, end_date, weekdays, holidays, count_holidays_as_workdays)
    total_days = (end_date - start_date).days + 1

    ws[f"A{current_row}"] = f"対象期間: {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}"
    ws[f"A{current_row}"].font = Font(bold=True)
    current_row += 1
    ws[f"A{current_row}"] = f"総日数: {total_days}日 / 営業日数: {total_workdays}日"
    ws[f"A{current_row}"].font = Font(bold=True)

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 5

    return wb


def parse_form_payload(form):
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start_default = today.strftime("%Y-%m-01")
    end_default = today.strftime("%Y-%m-%d")

    start_date_str = form.get("start_date", start_default)
    end_date_str = form.get("end_date", end_default)

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    except ValueError:
        start_date = datetime.strptime(start_default, "%Y-%m-%d")

    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
    except ValueError:
        end_date = datetime.strptime(end_default, "%Y-%m-%d")

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    selected_weekdays = form.getlist("weekdays[]")
    weekdays = {int(v) for v in selected_weekdays if v.isdigit() and 0 <= int(v) <= 6}
    if not weekdays:
        weekdays = {1, 2, 3, 4, 5}

    manual_holidays = parse_holidays(form.get("holidays", ""))
    include_japan_holidays = form.get("include_japan_holidays") == "on"
    holidays = set(manual_holidays)
    if include_japan_holidays:
        holidays.update(JAPAN_HOLIDAYS)

    count_holidays_as_workdays = form.get("count_holidays") == "on"

    preview_limit = 15
    try:
        preview_limit = max(5, min(60, int(form.get("preview_limit") or 15)))
    except ValueError:
        preview_limit = 15

    return {
        "start_date": start_date,
        "end_date": end_date,
        "weekdays": weekdays,
        "holidays": holidays,
        "manual_holidays": sorted(manual_holidays),
        "count_holidays_as_workdays": count_holidays_as_workdays,
        "include_japan_holidays": include_japan_holidays,
        "preview_limit": preview_limit,
    }


def build_default_form_state():
    now = datetime.now()
    start_of_month = datetime(now.year, now.month, 1)
    return {
        "start_date": start_of_month.strftime("%Y-%m-%d"),
        "end_date": now.strftime("%Y-%m-%d"),
        "weekdays": [1, 2, 3, 4, 5],
        "holidays": "",
        "count_holidays": False,
        "include_japan_holidays": True,
        "preview_limit": 15,
    }


def build_form_state(payload):
    return {
        "start_date": payload["start_date"].strftime("%Y-%m-%d"),
        "end_date": payload["end_date"].strftime("%Y-%m-%d"),
        "weekdays": sorted(list(payload["weekdays"])),
        "holidays": ",".join(payload["manual_holidays"]),
        "count_holidays": payload["count_holidays_as_workdays"],
        "include_japan_holidays": payload["include_japan_holidays"],
        "preview_limit": payload["preview_limit"],
    }


def build_result(payload):
    start_date = payload["start_date"]
    end_date = payload["end_date"]
    weekdays = payload["weekdays"]
    holidays = payload["holidays"]
    count_holidays_as_workdays = payload["count_holidays_as_workdays"]

    total_days = (end_date - start_date).days + 1
    workday_count = count_workdays(start_date, end_date, weekdays, holidays, count_holidays_as_workdays)
    non_workday_count = total_days - workday_count

    holiday_hits = sum(1 for day in each_day(start_date, end_date) if is_holiday(day, holidays))
    ratio = round((workday_count / total_days) * 100, 1) if total_days else 0

    workday_list = [
        day.strftime("%Y-%m-%d")
        for day in each_day(start_date, end_date)
        if is_workday(day, weekdays, holidays, count_holidays_as_workdays)
    ]

    preview = workday_list[: payload["preview_limit"]]

    return {
        "calendar": generate_calendar(start_date, end_date, weekdays, holidays, count_holidays_as_workdays),
        "stats": {
            "workday_count": workday_count,
            "non_workday_count": non_workday_count,
            "total_days": total_days,
            "holiday_hits": holiday_hits,
            "workday_ratio": ratio,
        },
        "workday_preview": preview,
        "workday_total_list_count": len(workday_list),
    }


@workday_bp.route("/", methods=["GET", "POST"])
@login_required
def workday():
    context = {
        "form_state": build_default_form_state(),
        "calendar": None,
        "stats": None,
        "workday_preview": [],
        "workday_total_list_count": 0,
        "weekday_labels": WEEKDAY_LABELS,
        "errors": [],
    }

    if request.method == "POST":
        payload = parse_form_payload(request.form)
        context["form_state"] = build_form_state(payload)

        period_length = (payload["end_date"] - payload["start_date"]).days
        if period_length > 366 * 3:
            context["errors"].append("期間は3年以内で指定してください。")
        else:
            result = build_result(payload)
            context.update(result)

    return render_template("workday.html", **context)


@workday_bp.route("/download", methods=["POST"])
@login_required
def download_excel():
    try:
        payload = parse_form_payload(request.form)
        wb = generate_excel_calendar(
            payload["start_date"],
            payload["end_date"],
            payload["weekdays"],
            payload["holidays"],
            payload["count_holidays_as_workdays"],
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"営業日カレンダー_{payload['start_date'].strftime('%Y%m%d')}_{payload['end_date'].strftime('%Y%m%d')}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception:
        return "エラーが発生しました", 500
