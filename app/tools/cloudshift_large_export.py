from __future__ import annotations

import io
from calendar import monthrange
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from xml.sax.saxutils import escape


def _safe_cell(value: Any) -> Any:
    text = str(value or "")
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _code_map(config: Mapping[str, Any]) -> dict[str, Mapping[str, Any]]:
    return {str(code.get("key") or "").casefold(): code for code in config.get("codes", []) if isinstance(code, Mapping)}


def _entry_maps(month_data: Mapping[str, Any]) -> dict[str, dict[str, Mapping[str, Any]]]:
    result: dict[str, dict[str, Mapping[str, Any]]] = {}
    for day, entries in (month_data.get("entries_per_day") or {}).items():
        result[str(day)] = {
            str(entry.get("member_id") or ""): entry
            for entry in entries if isinstance(entry, Mapping) and entry.get("member_id")
        }
    return result


def _entry_label(entry: Mapping[str, Any]) -> str:
    assignments = entry.get("assignments") if isinstance(entry.get("assignments"), list) else []
    if not assignments:
        return str(entry.get("value") or "")
    labels = []
    for item in assignments:
        if not isinstance(item, Mapping) or not item.get("code_key"):
            continue
        code = str(item.get("code_key") or "")
        if str(item.get("source_type") or "local") != "local":
            site = str(item.get("source_site_name") or item.get("source_project_title") or "他現場")
            custom = str(item.get("custom_label") or "")
            labels.append(f"{site}・{custom}" if custom else site)
        else:
            labels.append(code)
    return " / ".join(labels)


def _baseline_changed(day: str, member_id: str, entry: Mapping[str, Any], month_data: Mapping[str, Any]) -> bool:
    baseline = ((month_data.get("meta_data") or {}).get("baseline") or {}).get("entries_per_day") or {}
    previous = next((item for item in baseline.get(day, []) if str(item.get("member_id") or "") == member_id), {})
    def comparable(row: Mapping[str, Any], key: str) -> Any:
        if key in {"value", "holiday_kind"}:
            return str(row.get(key) or "")
        if key == "time_override":
            value = row.get(key)
            if isinstance(value, Mapping) and value.get("start") and value.get("end"):
                return {"start": str(value["start"]), "end": str(value["end"])}
            return None
        if key == "assignments":
            assignments = row.get("assignments") if isinstance(row.get("assignments"), list) else []
            if not assignments and row.get("value"):
                assignments = [{"code_key": row.get("value"), "source_type": "local"}]
            return [
                {
                    "code_key": str(item.get("code_key") or ""),
                    "source_type": str(item.get("source_type") or "local"),
                    "source_project_id": str(item.get("source_project_id") or ""),
                    "source_project_title": str(item.get("source_project_title") or ""),
                    "source_site_row_id": str(item.get("source_site_row_id") or ""),
                    "source_site_id": str(item.get("source_site_id") or ""),
                    "source_site_name": str(item.get("source_site_name") or ""),
                    "option_key": str(item.get("option_key") or ""),
                    "second_option": str(item.get("second_option") or ""),
                    "custom_label": str(item.get("custom_label") or ""),
                }
                for item in assignments if isinstance(item, Mapping) and item.get("code_key")
            ]
        value = row.get(key)
        return None if value in (None, "") else int(value)

    keys = ("value", "assignments", "holiday_kind", "time_override", "bind_override_minutes")
    return any(comparable(previous, key) != comparable(entry, key) for key in keys)


def _leave_counts_text(counts: Mapping[str, Any]) -> str:
    labels = {
        "rest": "休み", "rest_requested": "希望休",
        "substitute_rest": "振休", "substitute_rest_requested": "希望振休",
        "paid": "有休", "paid_requested": "希望有休",
        "special": "特休", "special_requested": "希望特休",
        "other": "その他", "other_requested": "希望その他", "empty": "未入力",
    }
    return " / ".join(
        f"{labels.get(key, key)}:{count}" for key, count in (counts or {}).items() if count
    )


def large_xlsx_bytes(
    title: str,
    config: Mapping[str, Any],
    month_data: Mapping[str, Any],
    worktime: Mapping[str, Any],
    *,
    highlight: bool = False,
) -> io.BytesIO:
    workbook = Workbook()
    shift = workbook.active
    shift.title = "シフト表"
    members = [member for member in config.get("members", []) if member.get("active", True)]
    entries = _entry_maps(month_data)
    codes = _code_map(config)
    meta = month_data.get("meta_data") or {}
    notes = meta.get("day_notes") or {}
    shift.append([_safe_cell(title), f"{month_data['year']}年{month_data['month']}月"])
    shift.append(["日付", "曜日", "日メモ", *[_safe_cell(member.get("display_name")) for member in members]])
    thin = Side(style="thin", color="C7D7E8")
    scheduled_side = Side(style="medium", color="2563EB")
    legal_side = Side(style="medium", color="DC2626")
    changed_fill = PatternFill("solid", fgColor="FFF3BF")
    weekdays = "月火水木金土日"
    for day in range(1, monthrange(int(month_data["year"]), int(month_data["month"]))[1] + 1):
        from datetime import date
        weekday = weekdays[date(int(month_data["year"]), int(month_data["month"]), day).weekday()]
        row = [day, weekday, _safe_cell(notes.get(str(day), ""))]
        for member in members:
            entry = entries.get(str(day), {}).get(str(member.get("id")), {})
            prefix = {"scheduled": "所", "legal": "法"}.get(str(entry.get("holiday_kind") or ""), "")
            changed = highlight and _baseline_changed(str(day), str(member.get("id")), entry, month_data)
            row.append(_safe_cell(("*" if changed else "") + prefix + _entry_label(entry)))
        shift.append(row)
        excel_row = shift.max_row
        for offset, member in enumerate(members, start=4):
            entry = entries.get(str(day), {}).get(str(member.get("id")), {})
            code = codes.get(str(entry.get("value") or "").casefold(), {})
            color = str(code.get("color") or "#FFFFFF").lstrip("#")
            if len(color) == 6:
                shift.cell(excel_row, offset).fill = PatternFill("solid", fgColor=color)
            if highlight and _baseline_changed(str(day), str(member.get("id")), entry, month_data):
                shift.cell(excel_row, offset).fill = changed_fill
            kind = str(entry.get("holiday_kind") or "")
            side = scheduled_side if kind == "scheduled" else legal_side if kind == "legal" else thin
            shift.cell(excel_row, offset).border = Border(left=side, right=side, top=side, bottom=side)
            if entry.get("comment"):
                from openpyxl.comments import Comment
                shift.cell(excel_row, offset).comment = Comment(str(entry["comment"]), "CloudShift")
    shift.freeze_panes = "D3"
    shift.auto_filter.ref = f"A2:{get_column_letter(3 + len(members))}{shift.max_row}"
    shift.column_dimensions["A"].width = 8
    shift.column_dimensions["B"].width = 6
    shift.column_dimensions["C"].width = 24
    for column in range(4, 4 + len(members)):
        shift.column_dimensions[get_column_letter(column)].width = 13
    for row in shift.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.border == Border():
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in shift[2]:
        cell.fill = PatternFill("solid", fgColor="DDEEFF")
        cell.font = Font(bold=True)

    summary = workbook.create_sheet("集計")
    summary.append(["メンバー", "拘束", "労働", "給与残業", "所定休出", "法定休出", "超過計", "安衛長時間", "勤務日数", "休み内訳", "最大連勤", "警告"])
    result_by_id = {str(person.get("person_id")): person for person in worktime.get("people", [])}
    for member in members:
        person = result_by_id.get(str(member.get("id")), {})
        totals = person.get("totals") or {}
        leave_counts = _leave_counts_text(totals.get("leave_counts") or {})
        summary.append([
            _safe_cell(member.get("display_name")), totals.get("bind_total_hhmm", "0:00"),
            totals.get("work_total_hhmm", "0:00"), totals.get("payroll_overtime_hhmm", "0:00"),
            totals.get("scheduled_holiday_work_hhmm", "0:00"), totals.get("legal_holiday_work_hhmm", "0:00"),
            totals.get("payroll_excess_total_hhmm", "0:00"), totals.get("anei_excess_hhmm", "0:00"),
            totals.get("work_days", 0), leave_counts, totals.get("max_consecutive_work_days", 0), len(person.get("violations") or []),
        ])
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    summary.column_dimensions["A"].width = 22
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="DDEEFF")
        cell.font = Font(bold=True)

    details = workbook.create_sheet("個人明細")
    for member in members:
        person = result_by_id.get(str(member.get("id")), {})
        details.append([_safe_cell(member.get("display_name"))])
        details.append(["日", "コード", "区分", "始業", "終業", "休憩", "拘束", "労働", "超過", "警告"])
        for day_result in person.get("days", []):
            if not day_result.get("code_key") and not day_result.get("warnings"):
                continue
            details.append([
                day_result.get("day"), _safe_cell(day_result.get("code_key")), day_result.get("category"),
                day_result.get("start") or "", day_result.get("end") or "", day_result.get("break_hhmm", "0:00"),
                day_result.get("bind_hhmm", "0:00"), day_result.get("work_hhmm", "0:00"),
                day_result.get("overtime_hhmm", "0:00"), ", ".join(day_result.get("warnings") or []),
            ])
        details.append([])
    details.freeze_panes = "A1"
    for column, width in enumerate((8, 16, 24, 10, 10, 10, 10, 10, 10, 28), start=1):
        details.column_dimensions[get_column_letter(column)].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


_PDF_FONT = "CloudShiftNotoJPSemiBold"


def _ensure_pdf_font() -> str:
    if _PDF_FONT not in pdfmetrics.getRegisteredFontNames():
        path = Path(__file__).resolve().parents[1] / "static" / "fonts" / "NotoSansJP-SemiBold.ttf"
        if not path.exists():
            raise RuntimeError("日本語PDFフォントが見つかりません")
        pdfmetrics.registerFont(TTFont(_PDF_FONT, str(path)))
    return _PDF_FONT


def _pdf_markup(value: Any) -> str:
    return escape(str(value or "")).replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br/>")


def _pdf_shift_cell(entry: Mapping[str, Any], label: str, style: ParagraphStyle) -> Paragraph:
    parts = [_pdf_markup(label)] if label else []
    comment = str(entry.get("comment") or "").strip()
    if comment:
        parts.append(f'<font size="7.5" color="#334155">※ {_pdf_markup(comment)}</font>')
    return Paragraph("<br/>".join(parts) or "　", style)


def large_pdf_bytes(
    title: str,
    config: Mapping[str, Any],
    month_data: Mapping[str, Any],
    worktime: Mapping[str, Any],
    *,
    highlight: bool = False,
) -> io.BytesIO:
    font = _ensure_pdf_font()
    members = [member for member in config.get("members", []) if member.get("active", True)]
    entries = _entry_maps(month_data)
    use_a3 = len(members) > 6
    page_size = landscape(A3 if use_a3 else A4)
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=page_size, leftMargin=8 * mm, rightMargin=8 * mm, topMargin=8 * mm, bottomMargin=8 * mm)
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = font
    cell_style = ParagraphStyle(
        "LargeShiftCell",
        parent=styles["BodyText"],
        fontName=font,
        fontSize=8.5,
        leading=10.5,
        alignment=1,
        textColor=colors.HexColor("#172033"),
        wordWrap="CJK",
        spaceBefore=0,
        spaceAfter=0,
    )
    header_style = ParagraphStyle(
        "LargeShiftHeader",
        parent=cell_style,
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#0F2742"),
    )
    note_style = ParagraphStyle(
        "LargeShiftNote",
        parent=cell_style,
        fontSize=8,
        leading=9.5,
        textColor=colors.HexColor("#334155"),
    )
    story: list[Any] = [Paragraph(f"{_safe_cell(title)}　{month_data['year']}年{month_data['month']}月", styles["Title"]), Spacer(1, 4 * mm)]
    notes = (month_data.get("meta_data") or {}).get("day_notes") or {}
    group_size = 10 if use_a3 else 8
    for group_index in range(0, max(1, len(members)), group_size):
        group = members[group_index:group_index + group_size]
        table_data = [[
            Paragraph("日", header_style),
            Paragraph("メモ", header_style),
            *[Paragraph(_pdf_markup(member.get("display_name")), header_style) for member in group],
        ]]
        for day in range(1, monthrange(int(month_data["year"]), int(month_data["month"]))[1] + 1):
            row = [Paragraph(str(day), header_style), Paragraph(_pdf_markup(notes.get(str(day), "")) or "　", note_style)]
            for member in group:
                entry = entries.get(str(day), {}).get(str(member.get("id")), {})
                marker = {"scheduled": "所", "legal": "法"}.get(str(entry.get("holiday_kind") or ""), "")
                changed = highlight and _baseline_changed(str(day), str(member.get("id")), entry, month_data)
                label = ("*" if changed else "") + marker + _entry_label(entry)
                row.append(_pdf_shift_cell(entry, label, cell_style))
            table_data.append(row)
        fixed_width = 42 * mm
        member_width = max(24 * mm, (document.width - fixed_width) / max(1, len(group)))
        widths = [10 * mm, 32 * mm] + [member_width for _ in group]
        table = Table(table_data, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DDEEFF")),
            ("GRID", (0, 0), (-1, -1), 0.65, colors.HexColor("#8EABC7")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
            ("TOPPADDING", (0, 0), (-1, 0), 4), ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
            ("TOPPADDING", (0, 1), (-1, -1), 3.2), ("BOTTOMPADDING", (0, 1), (-1, -1), 3.2),
            ("LEFTPADDING", (0, 0), (-1, -1), 2.5), ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
        ]))
        story.append(table)
        if group_index + group_size < len(members):
            story.append(PageBreak())
    story.extend([PageBreak(), Paragraph("月次集計", styles["Heading1"])])
    summary = [["メンバー", "拘束", "労働", "残業", "所定休出", "法定休出", "超過", "安衛長時間", "勤務日", "休み内訳", "警告"]]
    results = {str(person.get("person_id")): person for person in worktime.get("people", [])}
    for member in members:
        person = results.get(str(member.get("id")), {})
        totals = person.get("totals") or {}
        leave_counts = _leave_counts_text(totals.get("leave_counts") or {})
        summary.append([
            str(member.get("display_name") or ""), totals.get("bind_total_hhmm", "0:00"),
            totals.get("work_total_hhmm", "0:00"), totals.get("payroll_overtime_hhmm", "0:00"),
            totals.get("scheduled_holiday_work_hhmm", "0:00"), totals.get("legal_holiday_work_hhmm", "0:00"),
            totals.get("payroll_excess_total_hhmm", "0:00"), totals.get("anei_excess_hhmm", "0:00"),
            str(totals.get("work_days", 0)), leave_counts, str(len(person.get("violations") or [])),
        ])
    summary_table = Table(summary, repeatRows=1)
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DDEEFF")),
        ("GRID", (0, 0), (-1, -1), 0.65, colors.HexColor("#8EABC7")),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_table)
    document.build(story)
    output.seek(0)
    return output
