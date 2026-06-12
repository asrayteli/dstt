from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
import os
import csv
import io
from datetime import datetime
from openpyxl import load_workbook
import traceback

from app.models import Site, SiteBranch, SiteContractMaster, db
from app.site_contract_master import load_site_mapping, manager_ids_match

monthly_generator_bp = Blueprint("monthly_generator", __name__, url_prefix="/tools/monthly_generator")

# 経費対照表（システムに組み込み）
# 注：自動車売上は請負形態名称で判別するため、ここには含めない
EXPENSE_MAPPING = {
    "外注費": "材料",
    "消耗品費": "材料",
    "保健衛生費": "材料",
    "燃料費": "材料",
    "タイヤ費": "材料",
    "傭車費": "材料",
    "被服費": "材料",
    "修繕費": "材料",
    "賃借料": "材料",
    "保険料": "材料",
    "家賃地代": "材料",
    "減価償却費": "材料",
    "減価償却費（ﾘｰｽ資産）": "材料",
    "租税公課": "材料",
    "支払手数料": "材料",
    "材料その他": "材料",
    "給料": "労務",
    "賞与": "労務",
    "賞与引当金繰入額": "労務",
    "退職金": "労務",
    "退職給付費用": "労務",
    "法定福利費": "労務",
    "福利厚生費": "労務",
    "通勤費": "労務",
    "労務費振替": "労務",
    "労務費その他": "労務",
    "雇用調整助成金": "労務",
    "広告宣伝費": "経費",
    "募集費": "経費",
    "旅費交通費": "経費",
    "水道光熱費": "経費",
    "通信費": "経費",
    "リース料": "経費",
    "交際接待費": "経費",
    "会議費": "経費",
    "文化教育費": "経費",
    "事故負担金": "経費",
    "寄付金": "経費",
    "会費": "経費",
    "雑費": "経費",
    "経費その他": "経費",
}

# セル位置マッピング
CELL_MAPPING = {
    # 当期（単月）
    "材料": {"役員": "F9", "一般": "F17", "旅客": "F25", "役員合算": "N9", "一般合算": "N17", "旅客合算": "N25"},
    "労務": {"役員": "F10", "一般": "F18", "旅客": "F26", "役員合算": "N10", "一般合算": "N18", "旅客合算": "N26"},
    "経費": {"役員": "F11", "一般": "F19", "旅客": "F27", "役員合算": "N11", "一般合算": "N19", "旅客合算": "N27"},
    "基本売上": {"役員": "F6", "一般": "F14", "旅客": "F22", "役員合算": "N6", "一般合算": "N14", "旅客合算": "N22"},
    "その他売上": {"役員": "F7", "一般": "F15", "旅客": "F23", "役員合算": "N7", "一般合算": "N15", "旅客合算": "N23"},
}

# 前期セル位置マッピング
PREV_YEAR_CELL_MAPPING = {
    # 前期（単月）
    "材料": {"役員": "C9", "一般": "C17", "旅客": "C25", "役員合算": "L9", "一般合算": "L17", "旅客合算": "L25"},
    "労務": {"役員": "C10", "一般": "C18", "旅客": "C26", "役員合算": "L10", "一般合算": "L18", "旅客合算": "L26"},
    "経費": {"役員": "C11", "一般": "C19", "旅客": "C27", "役員合算": "L11", "一般合算": "L19", "旅客合算": "L27"},
    "基本売上": {"役員": "C6", "一般": "C14", "旅客": "C22", "役員合算": "L6", "一般合算": "L14", "旅客合算": "L22"},
    "その他売上": {"役員": "C7", "一般": "C15", "旅客": "C23", "役員合算": "L7", "一般合算": "L15", "旅客合算": "L23"},
}

# 見込セル位置マッピング（読み取り元）
FORECAST_SOURCE_MAPPING = {
    # 行番号のみ（列は対象月で決定: 4月=F, 5月=G, ..., 3月=Q）
    "基本売上": {"役員": 4, "一般": 15, "旅客": 26},
    "その他売上": {"役員": 5, "一般": 16, "旅客": 27},
    "材料": {"役員": 7, "一般": 18, "旅客": 29},
    "労務": {"役員": 8, "一般": 19, "旅客": 30},
    "経費": {"役員": 9, "一般": 20, "旅客": 31},
}

# 見込セル位置マッピング（書き込み先）
FORECAST_TARGET_MAPPING = {
    "基本売上": {"役員": "E6", "一般": "E14", "旅客": "E22"},
    "その他売上": {"役員": "E7", "一般": "E15", "旅客": "E23"},
    "材料": {"役員": "E9", "一般": "E17", "旅客": "E25"},
    "労務": {"役員": "E10", "一般": "E18", "旅客": "E26"},
    "経費": {"役員": "E11", "一般": "E19", "旅客": "E27"},
}

# 予算セル位置マッピング（書き込み先）
BUDGET_TARGET_MAPPING = {
    # 単月: D列
    "単月": {
        "基本売上": {"役員": "D6", "一般": "D14", "旅客": "D22"},
        "その他売上": {"役員": "D7", "一般": "D15", "旅客": "D23"},
        "材料": {"役員": "D9", "一般": "D17", "旅客": "D25"},
        "労務": {"役員": "D10", "一般": "D18", "旅客": "D26"},
        "経費": {"役員": "D11", "一般": "D19", "旅客": "D27"},
    },
    # 累計: M列
    "累計": {
        "基本売上": {"役員": "M6", "一般": "M14", "旅客": "M22"},
        "その他売上": {"役員": "M7", "一般": "M15", "旅客": "M23"},
        "材料": {"役員": "M9", "一般": "M17", "旅客": "M25"},
        "労務": {"役員": "M10", "一般": "M18", "旅客": "M26"},
        "経費": {"役員": "M11", "一般": "M19", "旅客": "M27"},
    }
}

def get_upload_folder():
    """一時アップロードフォルダのパスを取得"""
    folder = os.path.join(current_app.root_path, 'static', 'monthly_generator', 'uploads')
    os.makedirs(folder, exist_ok=True)
    return folder


def _normalize_site_id_for_master(value):
    text = str(value or '').strip()
    if not text or not text.isdigit():
        raise ValueError('契約番号が不正です')
    if len(text) > 5:
        raise ValueError('契約番号は5桁以内で指定してください')
    return text.zfill(5)


def _normalize_site_branch_for_master(value):
    text = str(value or '').strip()
    if not text or not text.isdigit():
        raise ValueError('枝番号が不正です')
    if len(text) > 3:
        raise ValueError('枝番号は3桁以内で指定してください')
    return text.zfill(3)


def _site_manager_name(site):
    return f"{site.site_manager_last} {site.site_manager_first}".strip()


def _upsert_segment_into_contract_master(contract_code, segment):
    master_row = db.session.get(SiteContractMaster, contract_code)
    if master_row is None:
        if len(contract_code) < 8 or not contract_code[:8].isdigit():
            return False
        site_id = _normalize_site_id_for_master(contract_code[:5])
        site_branch = _normalize_site_branch_for_master(contract_code[5:8])
        site = Site.query.filter_by(site_id=site_id).first()
        if site is None:
            return False
        branch = SiteBranch.query.filter_by(site_row_id=site.id, site_branch=site_branch).first()
        if branch is None:
            return False
        master_row = SiteContractMaster(
            contract_code=contract_code,
            site_row_id=site.id,
            site_branch_row_id=branch.id,
            site_id=site.site_id,
            site_branch=branch.site_branch,
            site_name=site.site_name,
            site_manager_id=site.site_manager_id,
            site_manager_name=_site_manager_name(site),
            cloudshift_option_key=branch.cloudshift_option_key,
            is_active=bool(site.is_active and branch.is_active),
            source='monthly_generator',
        )
        db.session.add(master_row)
    master_row.segment = segment
    if master_row.source in {'', 'siteplus'}:
        master_row.source = 'monthly_generator'
    return True


def _site_dict_from_contract_master(site_manager_id):
    manager_id = str(site_manager_id or '').strip()
    if not manager_id:
        raise ValueError('担当者IDを入力してください')
    site_dict, warnings, matched_rows = load_site_mapping(site_manager_id=manager_id)
    if matched_rows == 0:
        raise ValueError(f"担当者ID {manager_id} に一致する現場が現場リストPLUSに見つかりません")
    if matched_rows > 0 and not site_dict:
        raise ValueError(
            f"担当者ID {manager_id} の現場は見つかりましたが、セグメントが未登録です。"
            "現場リストPLUSで現場表を取り込んでください"
        )
    site_name_map = {}
    rows = (
        SiteContractMaster.query
        .filter(SiteContractMaster.is_active.is_(True))
        .order_by(SiteContractMaster.contract_code.asc())
        .all()
    )
    for row in rows:
        if not manager_ids_match(row.site_manager_id, manager_id):
            continue
        contract_code = str(row.contract_code or '').strip()
        if contract_code not in site_dict:
            continue
        site_name = str(row.site_name or '').strip()
        if site_name:
            site_name_map[contract_code] = site_name
    return site_dict, site_name_map, warnings


@monthly_generator_bp.route("/")
@login_required
def index():
    """メインページ"""
    return render_template("monthly_generator.html")


@monthly_generator_bp.route("/api/process", methods=["POST"])
@login_required
def process_files():
    """ファイル処理のメインエンドポイント"""
    try:
        # 必須ファイル取得
        subject_file = request.files.get('subject_file')  # 科目別推移表
        site_file = request.files.get('site_file')  # 現場表
        report_file = request.files.get('report_file')  # 月次報告書

        # オプションファイル取得
        prev_year_subject_file = request.files.get('prev_year_subject_file')  # 前年科目別推移表
        prev_year_site_file = request.files.get('prev_year_site_file')  # 前年現場表
        forecast_file = request.files.get('forecast_file')  # 見込ファイル
        budget_file = request.files.get('budget_file')  # 予算ファイル

        # パラメータ取得
        # 対象月（1-12）。未指定や非整数で 500 を出さないよう安全に検証する。
        try:
            target_month = int(request.form.get('target_month'))
        except (TypeError, ValueError):
            return jsonify({"error": "対象月（1〜12）を正しく指定してください"}), 400
        if not 1 <= target_month <= 12:
            return jsonify({"error": "対象月は1〜12で指定してください"}), 400
        sheet_name = request.form.get('sheet_name')  # シート名

        site_source = str(request.form.get('site_source', 'file') or 'file').strip().lower()
        site_manager_id = str(request.form.get('site_manager_id', '') or '').strip()
        ignore_missing_subject_sites = str(request.form.get('ignore_missing_subject_sites', '') or '').strip().lower() in {'1', 'true', 'yes', 'on'}

        if not all([subject_file, report_file, target_month, sheet_name]):
            return jsonify({"error": "必要なファイルまたはパラメータが不足しています"}), 400

        # ファイル保存
        upload_folder = get_upload_folder()
        if site_source not in {'file', 'db'}:
            return jsonify({"error": "site_source は file または db を指定してください"}), 400
        if site_source == 'file' and not site_file:
            return jsonify({"error": "現場表CSVを選択してください"}), 400
        if site_source == 'db' and not site_manager_id:
            return jsonify({"error": "担当者IDを入力してください"}), 400

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        user_id = current_user.username

        subject_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_subject.csv")
        site_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_site.csv") if site_source == 'file' else None
        report_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_report.xlsx")

        subject_file.save(subject_path)
        if site_path and site_file:
            site_file.save(site_path)
        report_file.save(report_path)

        # オプションファイル保存
        prev_year_subject_path = None
        prev_year_site_path = None
        forecast_path = None
        budget_path = None

        if prev_year_subject_file and prev_year_subject_file.filename:
            prev_year_subject_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_prev_subject.csv")
            prev_year_subject_file.save(prev_year_subject_path)

        if prev_year_site_file and prev_year_site_file.filename:
            prev_year_site_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_prev_site.csv")
            prev_year_site_file.save(prev_year_site_path)

        if forecast_file and forecast_file.filename:
            forecast_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_forecast.xlsx")
            forecast_file.save(forecast_path)

        if budget_file and budget_file.filename:
            budget_path = os.path.join(upload_folder, f"{user_id}_{timestamp}_budget.csv")
            budget_file.save(budget_path)

        # 処理実行
        result = process_monthly_data(
            subject_path, site_path, report_path, target_month, sheet_name,
            prev_year_subject_path, prev_year_site_path, forecast_path, budget_path,
            site_source=site_source,
            site_manager_id=site_manager_id,
            ignore_missing_subject_sites=ignore_missing_subject_sites
        )

        def _cleanup_input_files():
            try:
                os.remove(subject_path)
                if site_path:
                    os.remove(site_path)
                os.remove(report_path)
                if prev_year_subject_path:
                    os.remove(prev_year_subject_path)
                if prev_year_site_path:
                    os.remove(prev_year_site_path)
                if forecast_path:
                    os.remove(forecast_path)
                if budget_path:
                    os.remove(budget_path)
            except OSError:
                pass

        if result.get("confirmation_required"):
            db.session.rollback()
            _cleanup_input_files()
            return jsonify(result)

        if result.get("error"):
            db.session.rollback()
            _cleanup_input_files()
            return jsonify(result), 400
        db.session.commit()

        # 一時ファイル削除（入力ファイル全て）
        _cleanup_input_files()

        return jsonify(result)

    except Exception:
        db.session.rollback()
        # 詳細はサーバーログにのみ残し、クライアントには内部情報を返さない。
        current_app.logger.exception("monthly_generator process_files failed")
        return jsonify({"error": "処理中にエラーが発生しました。"}), 500


def process_monthly_data(subject_path, site_path, report_path, target_month, sheet_name,
                         prev_year_subject_path=None, prev_year_site_path=None, forecast_path=None, budget_path=None,
                         site_source='file', site_manager_id='', ignore_missing_subject_sites=False):
    """月次データ処理のメインロジック"""
    errors = []
    warnings = []  # オプションデータの警告用

    # ステップ1: CSVファイル読み込み
    try:
        # 科目別推移表読み込み（UTF-8, Shift-JIS, CP932を試行）
        subject_data = read_csv_with_encoding(subject_path)
        if not subject_data:
            return {"error": "科目別推移表の読み込みに失敗しました"}

        # 現場表読み込み
        site_data = read_csv_with_encoding(site_path) if site_source == 'file' else []
        if site_source == 'file' and not site_data:
            return {"error": "現場表の読み込みに失敗しました"}

        # 前年データ読み込み（オプション）
        prev_year_subject_data = None
        prev_year_site_data = None
        if prev_year_subject_path and prev_year_site_path:
            prev_year_subject_data = read_csv_with_encoding(prev_year_subject_path)
            prev_year_site_data = read_csv_with_encoding(prev_year_site_path)
            if not prev_year_subject_data or not prev_year_site_data:
                warnings.append("前年データの読み込みに失敗しました（スキップします）")
                prev_year_subject_data = None
                prev_year_site_data = None

    except Exception as e:
        return {"error": f"CSVファイルの読み込みエラー: {str(e)}"}

    # ステップ2: 現場表の解析（ヘッダー行をスキップ）
    # 新フォーマット: 契約コード,セグメント
    site_dict = {}  # 契約コードをキーとした辞書（8桁完全一致）
    site_name_map = {}

    if site_source == 'db':
        try:
            master_site_dict, master_site_name_map, master_warnings = _site_dict_from_contract_master(site_manager_id)
        except ValueError as e:
            return {"error": str(e)}
        if not master_site_dict:
            return {"error": "指定した担当者IDに紐づく有効な現場契約マスタがありません"}
        warnings.extend(master_warnings)
        site_name_map.update(master_site_name_map)
        site_data = [["contract_code", "segment"]] + [
            [contract_code, segment]
            for contract_code, segment in master_site_dict.items()
        ]

    for i, row in enumerate(site_data):
        if i == 0:  # ヘッダー行をスキップ
            continue
        if len(row) >= 2:
            contract_code = row[0].strip()
            segment = row[1].strip()

            # 契約コードが空の場合はスキップ
            if not contract_code:
                continue

            # セグメント値のバリデーション
            if segment not in ["役員", "一般", "旅客"]:
                warnings.append(f"不正なセグメント値: '{segment}' (契約コード: {contract_code}) - スキップします")
                continue

            # 8桁完全一致で辞書に登録（重複時は上書き）
            site_dict[contract_code] = segment
            if len(row) >= 3:
                site_name = row[2].strip()
                if site_name:
                    site_name_map[contract_code] = site_name

    # ステップ3: 科目別推移表から対象現場のデータを抽出（ヘッダー行をスキップ）
    extracted_data = []
    found_contracts = set()
    missing_contracts = set()  # 現場表に存在しない契約コード

    for i, row in enumerate(subject_data):
        if i == 0:  # ヘッダー行をスキップ
            continue
        if len(row) < 13:
            continue

        contract_code = row[8].strip() if len(row) > 8 else ""  # 契約コード
        corp_name = row[9].strip() if len(row) > 9 else ""
        site_name = row[10].strip() if len(row) > 10 else ""
        subject_code = row[11].strip() if len(row) > 11 else ""
        subject_name = row[12].strip() if len(row) > 12 else ""
        contract_type = row[5].strip() if len(row) > 5 else ""  # 請負形態名称

        # 科目コードが「販売費」の場合はスキップ
        if subject_code == "販売費":
            continue

        # 科目名称が空の場合は、間接原価以外はスキップ
        if not subject_name and subject_code != "間接原価":
            continue

        # 契約コードが空の場合はスキップ
        if not contract_code:
            continue

        # 現場表と照合（契約コードベース）
        segment = site_dict.get(contract_code)

        if segment:
            # 契約コードが見つかった
            found_contracts.add(contract_code)

            # 金額データ抽出（列13以降が各月のデータ、4月から開始）
            amounts = []
            for col_idx in range(13, len(row)):
                try:
                    amount = float(row[col_idx]) if row[col_idx] else 0
                    amounts.append(amount)
                except (ValueError, TypeError):
                    amounts.append(0)

            # 対象月に応じた必要データ数を計算
            if target_month >= 4:
                required_months = target_month - 4 + 1  # 4月選択→1, 5月選択→2, ..., 12月選択→9
            else:
                required_months = target_month + 8 + 1  # 1月選択→10, 2月選択→11, 3月選択→12

            # データ数チェック（警告のみ、処理は継続）
            if len(amounts) < required_months:
                warnings.append(f"データ不足: {contract_code} - {corp_name} (必要: {required_months}ヶ月分, 実際: {len(amounts)}ヶ月分)")

            extracted_data.append({
                "契約コード": contract_code,
                "法人名称": corp_name,
                "現場名称": site_name,
                "セグメント": segment,
                "科目名称": subject_name,
                "科目コード": subject_code,
                "請負形態名称": contract_type,
                "金額データ": amounts
            })
        else:
            # 現場表に存在しない契約コード（情報として記録するが、エラーにはしない）
            if contract_code:
                missing_contracts.add(f"{contract_code} ({corp_name} - {site_name})")

    # ステップ4: 現場表にあるが科目別推移表に見つからない現場の確認
    missing_subject_sites = []
    for contract_code in sorted(site_dict.keys()):
        if contract_code in found_contracts:
            continue
        missing_subject_sites.append({
            "contract_code": contract_code,
            "site_name": site_name_map.get(contract_code, "")
        })

    if missing_subject_sites and not ignore_missing_subject_sites:
        return {
            "confirmation_required": True,
            "message": "現場リストPLUSにはあるが、アップロードした科目別推移表に見つからない現場があります。これらを無視して処理を続けるか確認してください。",
            "missing_subject_sites": missing_subject_sites,
            "details": [
                f"契約コード: {item['contract_code']} / 現場名: {item['site_name'] or '不明'}"
                for item in missing_subject_sites
            ],
        }

    if missing_subject_sites:
        warnings.append(
            f"科目別推移表に見つからない現場 {len(missing_subject_sites)} 件を無視して処理を続行しました。"
        )

    # ステップ5: 経費ジャンル分けと合算
    aggregated = {
        "役員": {"材料": 0, "労務": 0, "経費": 0, "基本売上": 0, "その他売上": 0, "材料合算": 0, "労務合算": 0, "経費合算": 0, "基本売上合算": 0, "その他売上合算": 0},
        "一般": {"材料": 0, "労務": 0, "経費": 0, "基本売上": 0, "その他売上": 0, "材料合算": 0, "労務合算": 0, "経費合算": 0, "基本売上合算": 0, "その他売上合算": 0},
        "旅客": {"材料": 0, "労務": 0, "経費": 0, "基本売上": 0, "その他売上": 0, "材料合算": 0, "労務合算": 0, "経費合算": 0, "基本売上合算": 0, "その他売上合算": 0},
    }
    segment_contract_codes = {
        "役員": set(),
        "一般": set(),
        "旅客": set(),
    }

    unknown_subjects = set()

    for item in extracted_data:
        contract_code = item["契約コード"]
        segment = item["セグメント"]
        subject = item["科目名称"]
        amounts = item["金額データ"]
        subject_code = item["科目コード"]
        contract_type = item["請負形態名称"]

        if segment not in aggregated:
            continue
        segment_contract_codes[segment].add(contract_code)

        # 月のインデックス計算（4月が基準月、列13がamounts[0]）
        # 4-12月: month_index = target_month - 4 (0-8)
        # 1-3月: month_index = target_month + 8 (9-11)
        if target_month >= 4:
            month_index = target_month - 4
        else:
            month_index = target_month + 8

        month_amount = amounts[month_index] if 0 <= month_index < len(amounts) else 0

        # 累計金額（4月からtarget_monthまで）
        cumulative_amount = sum(amounts[:month_index + 1]) if 0 <= month_index < len(amounts) else 0

        # 間接原価の場合は労務に合算
        if subject_code == "間接原価":
            aggregated[segment]["労務"] += month_amount
            aggregated[segment]["労務合算"] += cumulative_amount
            continue

        # 自動車売上、または旅客セグメントの旅客運送売上の場合は請負形態名称で判別
        if subject == "自動車売上" or (subject == "旅客運送売上" and segment == "旅客"):
            if contract_type == "基本請負料":
                aggregated[segment]["基本売上"] += month_amount
                aggregated[segment]["基本売上合算"] += cumulative_amount
            elif contract_type == "その他請負料":
                aggregated[segment]["その他売上"] += month_amount
                aggregated[segment]["その他売上合算"] += cumulative_amount
            continue

        # 経費対照表から原価科目を取得
        expense_category = EXPENSE_MAPPING.get(subject)

        if not expense_category:
            unknown_subjects.add(subject)
            continue

        # 通常の経費
        aggregated[segment][expense_category] += month_amount
        aggregated[segment][f"{expense_category}合算"] += cumulative_amount

    # エラーチェック - 経費対照表に見つからない科目
    for subject in unknown_subjects:
        errors.append(f"科目が経費対照表に見つかりません: {subject}")

    # ステップ6: 前年データ処理（オプション）
    prev_year_aggregated = None
    if prev_year_subject_data and prev_year_site_data:
        try:
            prev_year_aggregated = process_prev_year_data(
                prev_year_subject_data, prev_year_site_data, target_month
            )
        except Exception as e:
            warnings.append(f"前年データ処理エラー: {str(e)}")

    # ステップ7: 見込データ抽出（オプション）
    forecast_data = None
    if forecast_path:
        forecast_data = extract_forecast_data(forecast_path, target_month, sheet_name)
        if forecast_data is None:
            warnings.append("見込ファイルの読み込みに失敗しました（スキップします）")

    # ステップ8: 予算データ抽出（オプション）
    budget_data = None
    if budget_path:
        budget_result = extract_budget_data(budget_path, target_month)
        if "error" in budget_result:
            warnings.append(f"予算ファイル処理エラー: {budget_result['error']}")
        else:
            budget_data = budget_result

    if errors:
        return {"error": "データ検証エラー", "details": errors}

    # ステップ8: Excelファイルへの書き込み
    try:
        wb = load_workbook(report_path)

        if sheet_name not in wb.sheetnames:
            return {"error": f"指定されたシート '{sheet_name}' が見つかりません"}

        ws = wb[sheet_name]

        # 当期データ書き込み
        for expense_type, cells in CELL_MAPPING.items():
            for segment in ["役員", "一般", "旅客"]:
                # 当月データ
                cell_addr = cells[segment]
                value = aggregated[segment][expense_type]
                ws[cell_addr] = value

                # 累計データ
                cell_addr_cumulative = cells[f"{segment}合算"]
                value_cumulative = aggregated[segment][f"{expense_type}合算"]
                ws[cell_addr_cumulative] = value_cumulative

        # 前年データ書き込み
        if prev_year_aggregated:
            for expense_type, cells in PREV_YEAR_CELL_MAPPING.items():
                for segment in ["役員", "一般", "旅客"]:
                    # 前期単月データ
                    cell_addr = cells[segment]
                    value = prev_year_aggregated[segment][expense_type]
                    ws[cell_addr] = value

                    # 前期累計データ
                    cell_addr_cumulative = cells[f"{segment}合算"]
                    value_cumulative = prev_year_aggregated[segment][f"{expense_type}合算"]
                    ws[cell_addr_cumulative] = value_cumulative

        # 見込データ書き込み
        if forecast_data:
            for expense_type, cells in FORECAST_TARGET_MAPPING.items():
                for segment in ["役員", "一般", "旅客"]:
                    cell_addr = cells[segment]
                    value = forecast_data[segment].get(expense_type, 0)
                    ws[cell_addr] = value

        # 予算データ書き込み
        if budget_data:
            # 単月データ書き込み
            for expense_type, cells in BUDGET_TARGET_MAPPING["単月"].items():
                for segment in ["役員", "一般", "旅客"]:
                    cell_addr = cells[segment]
                    value = budget_data["単月"][segment].get(expense_type, 0)
                    ws[cell_addr] = value

            # 累計データ書き込み
            for expense_type, cells in BUDGET_TARGET_MAPPING["累計"].items():
                for segment in ["役員", "一般", "旅客"]:
                    cell_addr = cells[segment]
                    value = budget_data["累計"][segment].get(expense_type, 0)
                    ws[cell_addr] = value

        # 保存
        output_path = report_path.replace('.xlsx', '_output.xlsx')
        wb.save(output_path)
        wb.close()

        if site_source == 'file':
            for contract_code, segment in site_dict.items():
                _upsert_segment_into_contract_master(contract_code, segment)

        # ファイル名のみを抽出
        output_filename = os.path.basename(output_path)

        result = {
            "success": True,
            "output_file": output_filename,
            "data": aggregated,
            "prev_year_data": prev_year_aggregated,  # 前期データ追加
            "forecast_data": forecast_data,  # 見込データ追加
            "budget_data": budget_data,  # 予算データ追加
            "debug": {
                "extracted_count": len(extracted_data),
                "found_contracts_count": len(found_contracts),
                "site_dict_count": len(site_dict),
                "missing_contracts": list(missing_contracts) if missing_contracts else [],
                "missing_subject_sites": missing_subject_sites,
                "segment_contract_codes": {
                    segment: sorted(codes)
                    for segment, codes in segment_contract_codes.items()
                },
            }
        }

        # 警告がある場合は結果に含める
        if warnings:
            result["warnings"] = warnings

        return result

    except Exception as e:
        traceback.print_exc()
        return {"error": f"Excelファイルの処理エラー: {str(e)}"}


def read_csv_with_encoding(file_path):
    """複数のエンコーディングでCSVを読み込む。

    utf-8-sig を最初に試すことで、BOM 付き UTF-8 の CSV を 'utf-8' で
    読んでしまい先頭セルに \\ufeff が残る不具合を防ぐ。
    """
    encodings = ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']

    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                data = list(reader)
                return data
        except (UnicodeDecodeError, UnicodeError):
            continue

    return None


def extract_forecast_data(forecast_path, target_month, sheet_name):
    """見込ファイルからデータを抽出する

    Args:
        forecast_path: 見込ファイルのパス
        target_month: 対象月（1-12）
        sheet_name: シート名

    Returns:
        dict: セグメントごとの見込データ
    """
    try:
        wb = load_workbook(forecast_path, data_only=True)

        if sheet_name not in wb.sheetnames:
            return None

        ws = wb[sheet_name]

        # 対象月から列を決定（4月=F=6列目, 5月=G=7列目, ..., 3月=Q=17列目）
        # 4月=4, 5月=5, ..., 12月=12, 1月=1, 2月=2, 3月=3
        if target_month >= 4:
            col_index = target_month - 4 + 6  # 4月→6, 5月→7, ..., 12月→14
        else:
            col_index = target_month + 8 + 6  # 1月→15, 2月→16, 3月→17

        # 列文字を取得
        col_letter = chr(ord('A') + col_index - 1)

        forecast_data = {
            "役員": {},
            "一般": {},
            "旅客": {}
        }

        # 各項目のデータを抽出
        for expense_type, rows in FORECAST_SOURCE_MAPPING.items():
            for segment, row_num in rows.items():
                cell_addr = f"{col_letter}{row_num}"
                value = ws[cell_addr].value

                # 空白の場合は0
                if value is None or value == "":
                    value = 0
                else:
                    try:
                        value = float(value)
                    except Exception:
                        value = 0

                forecast_data[segment][expense_type] = value

        wb.close()
        return forecast_data

    except Exception as e:
        traceback.print_exc()
        return None


def extract_budget_data(budget_path, target_month):
    """予算ファイル（CSV）からデータを抽出する

    CSVフォーマット:
    - 1-3行目: ヘッダー情報
    - 4行目: 列ヘッダー（セグメント,項目,4月,5月,...,3月,累計）
    - 5-9行目: 役員データ（基本売上、その他売上、材料原価、労務間接、経費原価）
    - 10行目: 空行
    - 11-15行目: 一般データ
    - 16行目: 空行
    - 17-21行目: 旅客データ

    Args:
        budget_path: 予算CSVファイルのパス
        target_month: 対象月（1-12）

    Returns:
        dict: {
            "単月": {"役員": {...}, "一般": {...}, "旅客": {...}},
            "累計": {"役員": {...}, "一般": {...}, "旅客": {...}}
        }
        エラー時: {"error": "エラーメッセージ"}
    """
    try:
        # CSVファイル読み込み
        csv_data = read_csv_with_encoding(budget_path)
        if not csv_data:
            return {"error": "CSVファイルの読み込みに失敗しました（文字コードが不正です）"}

        # 最低限の行数チェック（ヘッダー4行 + 役員5行 + 空行1行 + 一般5行 + 空行1行 + 旅客5行 = 21行）
        if len(csv_data) < 21:
            return {"error": f"CSVファイルの行数が不足しています（期待: 21行以上、実際: {len(csv_data)}行）"}

        # ヘッダー行（4行目、インデックス3）から列位置を取得
        header_row = csv_data[3]
        if len(header_row) < 15:  # セグメント,項目,4-3月(12ヶ月),累計 = 最低15列
            return {"error": f"CSVヘッダーの列数が不足しています（期待: 15列以上、実際: {len(header_row)}列）"}

        # 対象月の列インデックスを計算（4月=列2、5月=列3、...、3月=列13）
        # target_month: 4=4月, 5=5月, ..., 12=12月, 1=1月, 2=2月, 3=3月
        if target_month >= 4:
            month_col_index = target_month - 4 + 2  # 4月→2, 5月→3, ..., 12月→10
        else:
            month_col_index = target_month + 8 + 2  # 1月→11, 2月→12, 3月→13

        # 項目名マッピング（CSVの項目名 → システムの項目名）
        item_mapping = {
            "基本売上": "基本売上",
            "その他売上": "その他売上",
            "材料原価": "材料",
            "労務間接": "労務",
            "経費原価": "経費"
        }

        # データ構造初期化
        budget_data = {
            "単月": {
                "役員": {},
                "一般": {},
                "旅客": {}
            },
            "累計": {
                "役員": {},
                "一般": {},
                "旅客": {}
            }
        }

        # セグメントごとのデータ抽出
        # 役員: 5-9行目（インデックス4-8）
        # 一般: 11-15行目（インデックス10-14）
        # 旅客: 17-21行目（インデックス16-20）
        segment_rows = {
            "役員": (4, 9),   # 5行目から9行目まで
            "一般": (10, 15),  # 11行目から15行目まで
            "旅客": (16, 21)   # 17行目から21行目まで
        }

        for segment, (start_idx, end_idx) in segment_rows.items():
            for row_idx in range(start_idx, end_idx):
                if row_idx >= len(csv_data):
                    return {"error": f"{segment}セグメントのデータ行が見つかりません（行{row_idx + 1}）"}

                row = csv_data[row_idx]
                if len(row) < 2:
                    continue  # 空行をスキップ

                # 項目名を取得（列1）
                csv_item_name = row[1].strip() if len(row) > 1 else ""
                if not csv_item_name:
                    continue

                # 項目名をマッピング
                system_item_name = item_mapping.get(csv_item_name)
                if not system_item_name:
                    return {"error": f"不明な項目名: '{csv_item_name}' (行{row_idx + 1})"}

                # 単月データ取得
                if month_col_index >= len(row):
                    return {"error": f"{segment}セグメントの{csv_item_name}の対象月データが見つかりません（行{row_idx + 1}、列{month_col_index + 1}）"}

                month_value_str = row[month_col_index].strip() if row[month_col_index] else "0"
                try:
                    month_value = float(month_value_str) if month_value_str else 0
                except ValueError:
                    return {"error": f"{segment}セグメントの{csv_item_name}の単月データが数値ではありません: '{month_value_str}' (行{row_idx + 1})"}

                budget_data["単月"][segment][system_item_name] = month_value

                # 累計データを計算（4月から対象月まで）
                cumulative_value = 0
                # 4月から対象月までループ
                if target_month >= 4:
                    # 4月から対象月まで（同じ年度内）
                    for month in range(4, target_month + 1):
                        col_idx = month - 4 + 2  # 4月→2, 5月→3, ...
                        if col_idx >= len(row):
                            return {"error": f"{segment}セグメントの{csv_item_name}の{month}月データが見つかりません（行{row_idx + 1}）"}
                        value_str = row[col_idx].strip() if row[col_idx] else "0"
                        try:
                            value = float(value_str) if value_str else 0
                            cumulative_value += value
                        except ValueError:
                            return {"error": f"{segment}セグメントの{csv_item_name}の{month}月データが数値ではありません: '{value_str}' (行{row_idx + 1})"}
                else:
                    # 4月から12月まで + 1月から対象月まで（年度をまたぐ）
                    # 4月から12月
                    for month in range(4, 13):
                        col_idx = month - 4 + 2
                        if col_idx >= len(row):
                            return {"error": f"{segment}セグメントの{csv_item_name}の{month}月データが見つかりません（行{row_idx + 1}）"}
                        value_str = row[col_idx].strip() if row[col_idx] else "0"
                        try:
                            value = float(value_str) if value_str else 0
                            cumulative_value += value
                        except ValueError:
                            return {"error": f"{segment}セグメントの{csv_item_name}の{month}月データが数値ではありません: '{value_str}' (行{row_idx + 1})"}
                    # 1月から対象月
                    for month in range(1, target_month + 1):
                        col_idx = month + 8 + 2  # 1月→11, 2月→12, 3月→13
                        if col_idx >= len(row):
                            return {"error": f"{segment}セグメントの{csv_item_name}の{month}月データが見つかりません（行{row_idx + 1}）"}
                        value_str = row[col_idx].strip() if row[col_idx] else "0"
                        try:
                            value = float(value_str) if value_str else 0
                            cumulative_value += value
                        except ValueError:
                            return {"error": f"{segment}セグメントの{csv_item_name}の{month}月データが数値ではありません: '{value_str}' (行{row_idx + 1})"}

                budget_data["累計"][segment][system_item_name] = cumulative_value

        return budget_data

    except Exception as e:
        traceback.print_exc()
        return {"error": f"予算ファイル処理中に予期しないエラーが発生しました: {str(e)}"}


def process_prev_year_data(subject_data, site_data, target_month):
    """前年データを処理する（現在のデータ処理と同じロジック）

    Args:
        subject_data: 前年科目別推移表データ
        site_data: 前年現場表データ
        target_month: 対象月（1-12）

    Returns:
        dict: セグメントごとの集計データ
    """
    # 現場表の解析
    site_dict = {}
    for i, row in enumerate(site_data):
        if i == 0:
            continue
        if len(row) >= 2:
            contract_code = row[0].strip()
            segment = row[1].strip()
            if not contract_code:
                continue
            site_dict[contract_code] = segment

    # データ抽出
    extracted_data = []
    for i, row in enumerate(subject_data):
        if i == 0:
            continue
        if len(row) < 13:
            continue

        contract_code = row[8].strip() if len(row) > 8 else ""
        subject_code = row[11].strip() if len(row) > 11 else ""
        subject_name = row[12].strip() if len(row) > 12 else ""
        contract_type = row[5].strip() if len(row) > 5 else ""

        if subject_code == "販売費":
            continue
        if not subject_name and subject_code != "間接原価":
            continue
        if not contract_code:
            continue

        segment = site_dict.get(contract_code)
        if segment:
            amounts = []
            for col_idx in range(13, len(row)):
                try:
                    amount = float(row[col_idx]) if row[col_idx] else 0
                    amounts.append(amount)
                except (ValueError, TypeError):
                    amounts.append(0)

            extracted_data.append({
                "セグメント": segment,
                "科目名称": subject_name,
                "科目コード": subject_code,
                "請負形態名称": contract_type,
                "金額データ": amounts
            })

    # 集計
    aggregated = {
        "役員": {"材料": 0, "労務": 0, "経費": 0, "基本売上": 0, "その他売上": 0, "材料合算": 0, "労務合算": 0, "経費合算": 0, "基本売上合算": 0, "その他売上合算": 0},
        "一般": {"材料": 0, "労務": 0, "経費": 0, "基本売上": 0, "その他売上": 0, "材料合算": 0, "労務合算": 0, "経費合算": 0, "基本売上合算": 0, "その他売上合算": 0},
        "旅客": {"材料": 0, "労務": 0, "経費": 0, "基本売上": 0, "その他売上": 0, "材料合算": 0, "労務合算": 0, "経費合算": 0, "基本売上合算": 0, "その他売上合算": 0},
    }

    for item in extracted_data:
        segment = item["セグメント"]
        subject = item["科目名称"]
        amounts = item["金額データ"]
        subject_code = item["科目コード"]
        contract_type = item["請負形態名称"]

        if segment not in aggregated:
            continue

        # 月のインデックス計算（4月が基準月）
        # 4-12月: month_index = target_month - 4 (0-8)
        # 1-3月: month_index = target_month + 8 (9-11)
        if target_month >= 4:
            month_index = target_month - 4
        else:
            month_index = target_month + 8

        month_amount = amounts[month_index] if 0 <= month_index < len(amounts) else 0
        cumulative_amount = sum(amounts[:month_index + 1]) if 0 <= month_index < len(amounts) else 0

        if subject_code == "間接原価":
            aggregated[segment]["労務"] += month_amount
            aggregated[segment]["労務合算"] += cumulative_amount
            continue

        if subject == "自動車売上" or (subject == "旅客運送売上" and segment == "旅客"):
            if contract_type == "基本請負料":
                aggregated[segment]["基本売上"] += month_amount
                aggregated[segment]["基本売上合算"] += cumulative_amount
            elif contract_type == "その他請負料":
                aggregated[segment]["その他売上"] += month_amount
                aggregated[segment]["その他売上合算"] += cumulative_amount
            continue

        expense_category = EXPENSE_MAPPING.get(subject)
        if expense_category:
            aggregated[segment][expense_category] += month_amount
            aggregated[segment][f"{expense_category}合算"] += cumulative_amount

    return aggregated


@monthly_generator_bp.route("/api/download/<filename>")
@login_required
def download_file(filename):
    """処理済みファイルのダウンロード"""
    try:
        upload_folder = get_upload_folder()
        file_path = os.path.join(upload_folder, filename)

        if not os.path.exists(file_path):
            return jsonify({"error": "ファイルが見つかりません"}), 404

        # ファイルをメモリに読み込む
        with open(file_path, 'rb') as f:
            file_data = io.BytesIO(f.read())

        # ファイルを削除（メモリに読み込んだ後なので削除可能）
        try:
            os.remove(file_path)
        except Exception as e:
            print(f"一時ファイル削除エラー: {e}")

        # メモリ上のデータを送信
        file_data.seek(0)
        return send_file(
            file_data,
            as_attachment=True,
            download_name="月次報告書_出力.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500
