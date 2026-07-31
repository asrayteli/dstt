import json
import subprocess
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.services.cloudshift_large import (
    calculate_large_month,
    default_large_config,
    effective_holiday_kind,
    effective_holiday_rule,
    holiday_class_for_day,
    normalize_large_config,
)
from app.tools.cloudshift_large_export import _baseline_changed, large_pdf_bytes, large_xlsx_bytes
from app.tools.shiftersync_format import normalize_large_entries_for_month


def _sample():
    config = default_large_config()
    config["members"] = [{"id": "m1", "display_name": "職員A", "active": True, "order": 10}]
    config["codes"].append({
        "key": "A", "label": "A", "category": "work", "active": True, "order": 1,
        "times": {key: {"start": "09:00", "end": "18:00"} for key in ("weekday", "saturday", "holiday")},
        "color": "#dbeafe",
    })
    config = normalize_large_config(config)
    month = {
        "year": 2026, "month": 7,
        "entries_per_day": {"1": [{"member_id": "m1", "value": "A", "employee_name": "職員A"}]},
        "meta_data": {"day_types": {}, "day_notes": {"1": "朝礼"}},
    }
    return config, month


def test_large_normalizer_deduplicates_and_discards_empty_cells():
    normalized = normalize_large_entries_for_month({
        "1": [
            {"member_id": "m1", "value": "A"},
            {"member_id": "m1", "value": "B", "comment": "後勝ち"},
            {"member_id": "m2", "value": ""},
            {"member_id": "m3", "value": "", "holiday_kind": "legal"},
        ]
    }, 2026, 7)
    assert [(row["member_id"], row["value"]) for row in normalized["1"]] == [("m1", "B"), ("m3", "")]
    assert normalized["1"][0]["comment"] == "後勝ち"


def test_large_config_recovers_legacy_substitute_column_type_from_id():
    config = default_large_config()
    config["members"] = [
        {
            "id": "mem_regular_1",
            "display_name": "Regular",
            "employee_number": "",
            "employee_name": "",
        },
        {
            "id": "sub_legacy_1",
            "display_name": "Substitute",
            "employee_number": "",
            "employee_name": "",
        },
    ]

    members = normalize_large_config(config)["members"]

    assert [member["column_type"] for member in members] == ["regular", "substitute"]


def test_default_leave_choices_and_external_assignment_metadata_are_user_facing():
    config = default_large_config()
    assert [code["label"] for code in config["codes"]] == [
        "法定休", "所定休", "振休", "有休", "希望休", "希望有休", "希望振休",
    ]
    normalized = normalize_large_entries_for_month({"1": [{
        "member_id": "m1",
        "assignments": [
            {"code_key": "A", "source_type": "local"},
            {"code_key": "E", "source_type": "scene", "source_project_id": "scene-2", "source_project_title": "別現場", "source_site_row_id": "22", "source_site_id": "site-2", "source_site_name": "第二現場", "option_key": "E", "second_option": "SUB", "custom_label": "応援"},
        ],
    }]}, 2026, 7)["1"][0]
    assert normalized["value"] == "A"
    assert [item["code_key"] for item in normalized["assignments"]] == ["A", "E"]
    assert normalized["assignments"][1] == {
        **normalized["assignments"][1],
        "source_type": "scene", "source_project_id": "scene-2",
        "source_project_title": "別現場", "source_site_row_id": "22",
        "source_site_id": "site-2", "source_site_name": "第二現場",
        "option_key": "E", "second_option": "SUB", "custom_label": "応援",
    }


def test_external_only_assignment_is_not_counted_as_empty_or_local_work():
    config, month = _sample()
    month["entries_per_day"] = {"1": [{
        "member_id": "m1", "value": "E",
        "assignments": [{"code_key": "E", "source_type": "scene", "source_site_name": "第二現場", "custom_label": "応援"}],
    }]}
    person = calculate_large_month(config, month, [])["people"][0]
    assert person["days"][0]["category"] == "external"
    assert person["totals"]["work_days"] == 0
    assert person["totals"]["leave_counts"]["empty"] == 30


def test_baseline_comparison_ignores_comment_only_changes():
    month = {"meta_data": {"baseline": {"entries_per_day": {"1": [
        {"member_id": "m1", "value": "A", "comment": "旧コメント"}
    ]}}}}
    assert not _baseline_changed("1", "m1", {"member_id": "m1", "value": "A", "comment": "新コメント"}, month)
    assert not _baseline_changed("2", "m1", {"member_id": "m1", "value": "", "comment": "コメントのみ"}, month)
    assert _baseline_changed("1", "m1", {"member_id": "m1", "value": "B", "comment": "旧コメント"}, month)


def test_large_exports_are_readable_and_contain_expected_sheets():
    config, month = _sample()
    month["entries_per_day"]["1"][0]["comment"] = "引継ぎあり"
    result = calculate_large_month(config, month, [])
    xlsx = large_xlsx_bytes("大規模テスト", config, month, result)
    workbook = load_workbook(xlsx, read_only=True)
    assert workbook.sheetnames == ["シフト表", "集計", "個人明細"]
    assert workbook["シフト表"]["D3"].value == "A"
    pdf = large_pdf_bytes("大規模テスト", config, month, result)
    reader = PdfReader(pdf)
    assert len(reader.pages) >= 1
    assert len(pdf.getvalue()) > 1000
    assert "引継ぎあり" in "".join(page.extract_text() or "" for page in reader.pages)


def test_holiday_rule_normalizes_weekdays_dates_and_member_override():
    config = default_large_config()
    config["settings"]["holiday_rule"] = {
        "legal_weekdays": [0, 7, "1", 0],
        "scheduled_weekdays": [6],
        "legal_dates": ["2026-08-11", "2026-08-11"],
        "scheduled_dates": ["2026-12-31"],
    }
    config["members"] = [
        {"id": "m1", "display_name": "共通"},
        {"id": "m2", "display_name": "個別", "holiday_rule": {"mode": "custom", "legal_weekdays": [3]}},
    ]
    normalized = normalize_large_config(config)
    rule = normalized["settings"]["holiday_rule"]
    # 0〜6 以外は捨て、重複は畳んで昇順に並べる。
    assert rule["legal_weekdays"] == [0, 1]
    assert rule["legal_dates"] == ["2026-08-11"]
    assert normalized["members"][0]["holiday_rule"]["mode"] == "default"
    assert normalized["members"][1]["holiday_rule"]["mode"] == "custom"

    settings = normalized["settings"]
    shared = effective_holiday_rule(normalized["members"][0], settings)
    custom = effective_holiday_rule(normalized["members"][1], settings)
    # 2026-08-11(火) は共通ルールでは指定日の法定休日、個別ルールでは所定労働日。
    assert holiday_class_for_day(shared, 2026, 8, 11) == "legal"
    assert holiday_class_for_day(custom, 2026, 8, 11) == ""
    # 2026-08-12(水) は個別ルールの曜日指定で法定休日。
    assert holiday_class_for_day(custom, 2026, 8, 12) == "legal"
    # セル側の上書きは曜日・指定日より強い。
    assert effective_holiday_kind(shared, 2026, 8, 11, {"holiday_kind": "none"}) == ""
    assert effective_holiday_kind(custom, 2026, 8, 11, {"holiday_kind": "scheduled"}) == "scheduled"


def test_invalid_holiday_rule_date_is_rejected():
    config = default_large_config()
    config["settings"]["holiday_rule"] = {"legal_dates": ["2026/08/11"]}
    with pytest.raises(ValueError):
        normalize_large_config(config)


def test_server_and_browser_agree_on_rule_based_holiday_totals():
    """休日ルール・休憩上書き・有休換算をサーバー計算とブラウザ計算で突き合わせる。"""
    config = default_large_config()
    config["settings"]["holiday_rule"] = {
        "legal_weekdays": [0], "scheduled_weekdays": [6],
        "legal_dates": [], "scheduled_dates": ["2026-07-02"],
    }
    config["members"] = [
        {"id": "m1", "display_name": "共通", "active": True, "order": 10},
        {"id": "m2", "display_name": "個別", "active": True, "order": 20,
         "holiday_rule": {"mode": "custom", "legal_weekdays": [3], "scheduled_weekdays": [4],
                          "legal_dates": [], "scheduled_dates": []}},
    ]
    config["codes"].append({
        "key": "A", "label": "A", "category": "work", "active": True, "order": 1,
        "times": {key: {"start": "08:00", "end": "18:00"} for key in ("weekday", "saturday", "holiday")},
        "color": "#dbeafe",
    })
    config = normalize_large_config(config)
    entries = {}
    for day in (1, 2, 3, 4, 5):
        entries[str(day)] = [
            {"member_id": member_id, "value": "A", "assignments": [{"code_key": "A", "source_type": "local"}],
             **({"break_override_minutes": 30} if day == 1 else {})}
            for member_id in ("m1", "m2")
        ]
    entries["6"] = [{"member_id": member_id, "value": "有休",
                     "assignments": [{"code_key": "有休", "source_type": "local"}]}
                    for member_id in ("m1", "m2")]
    month = {"year": 2026, "month": 7, "entries_per_day": entries, "meta_data": {"day_types": {}, "day_notes": {}}}

    server = calculate_large_month(config, month, [])
    script = (
        "const c=require('./app/static/cloudshift/js/cloudshift_large_calc.js');"
        " process.stdout.write(JSON.stringify(c.calculate(JSON.parse(process.argv[1]))));"
    )
    completed = subprocess.run(
        ["node", "-e", script, json.dumps(
            {"year": 2026, "month": 7, "config": config, "entries_per_day": entries, "holidays": []},
            ensure_ascii=False,
        )],
        cwd=Path(__file__).resolve().parents[1], text=True, encoding="utf-8", capture_output=True, check=True,
    )
    browser = json.loads(completed.stdout)

    compared = (
        "payroll_overtime_minutes", "scheduled_holiday_work_minutes", "legal_holiday_work_minutes",
        "agreement36_minutes", "special_clause_minutes", "annual960_minutes",
        "work_total_minutes", "break_total_minutes", "paid_leave_credit_minutes",
        "work_with_paid_total_minutes",
    )
    for index in range(2):
        server_totals = server["people"][index]["totals"]
        browser_totals = browser["people"][index]["totals"]
        assert {key: server_totals[key] for key in compared} == {key: browser_totals[key] for key in compared}
        assert [day["category"] for day in server["people"][index]["days"]] == \
            [day["category"] for day in browser["people"][index]["days"]]

    # 共通ルール: 7/2(木)=指定日の所定休 / 7/4(土)=所定休 / 7/5(日)=法定休。
    shared_days = {day["day"]: day["category"] for day in server["people"][0]["days"] if day["category"] != "empty"}
    assert shared_days == {
        1: "work", 2: "scheduled_holiday_work", 3: "work",
        4: "scheduled_holiday_work", 5: "legal_holiday_work", 6: "leave",
    }
    # 個別ルール: 水=法定休 / 木=所定休。土日は所定労働日。
    custom_days = {day["day"]: day["category"] for day in server["people"][1]["days"] if day["category"] != "empty"}
    assert custom_days == {
        1: "legal_holiday_work", 2: "scheduled_holiday_work", 3: "work",
        4: "work", 5: "work", 6: "leave",
    }
