"""CSV / XLSX エクスポートの数式インジェクション対策の回帰テスト。

シフト名・タイトル・コメント等は利用者入力（公開編集URL経由でも入る）であり、
先頭が =, +, -, @ のセルをそのまま書き出すと、Excel 等で開いた際に数式として
評価されてしまう（=HYPERLINK(...) など）。エクスポート時に ' を前置して
無害化することを保証する。あわせて、通常データの CSV ラウンドトリップ
（エクスポート→取込）が壊れていないことも確認する。
"""

import importlib.util
import io
import sys
from pathlib import Path
from types import SimpleNamespace

from flask import Flask
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.models import db


def _load_module():
    spec = importlib.util.spec_from_file_location(
        "cloudshift_formula_guard_module", ROOT / "app" / "tools" / "cloudshift.py"
    )
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


def _build_client(tmp_path):
    module = _load_module()
    app = Flask(__name__, root_path=str(tmp_path), instance_path=str(tmp_path / "instance"))
    app.secret_key = "test"
    app.config["TESTING"] = True
    app.config["LOGIN_DISABLED"] = True
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{(tmp_path / 'guard.db').as_posix()}"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(app)
    app.register_blueprint(module.cloudshift_bp)
    with app.app_context():
        db.create_all()
    return module, app.test_client()


def _owner():
    return SimpleNamespace(is_authenticated=True, username="owner01", name="Owner")


BASE = "/tools/shiftersync/cloudshift"


def _create_with_payload(client, title, entries):
    resp = client.post(
        f"{BASE}/api/create",
        data={"title": title, "mode": "scene", "year": "2026", "month": "4", "required_capacity": "1"},
    )
    pid = resp.get_json()["project"]["project"]["id"]
    client.put(
        f"{BASE}/api/project/{pid}/month/2026/4",
        json={
            "base_month": {"year": 2026, "month": 4, "required_capacity": 1, "entries_per_day": {}},
            "required_capacity": 1,
            "entries_per_day": entries,
        },
    )
    return pid


def test_neutralizes_csv_and_xlsx_formula_injection(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _owner()

    entries = {
        "1": [
            {"id": "e1", "value": '=HYPERLINK("http://evil","x")', "comment": "+danger"},
            {"id": "e2", "value": "@SUM(A1)"},
            {"id": "e3", "value": "-2+3"},
            {"id": "e4", "value": "田中"},
        ]
    }
    pid = _create_with_payload(client, "=1+1", entries)

    # --- CSV ---
    csv_text = client.get(f"{BASE}/api/project/{pid}/export/csv?month_key=2026-04").data.decode("utf-8-sig")
    assert "'=HYPERLINK" in csv_text
    assert "'@SUM(A1)" in csv_text
    assert "'-2+3" in csv_text
    assert "'+danger" in csv_text
    assert "'=1+1" in csv_text  # タイトルも無害化
    # 通常値は ' を付けない。
    assert "田中" in csv_text and "'田中" not in csv_text

    # --- XLSX ---
    xlsx = client.get(f"{BASE}/api/project/{pid}/export/xlsx?month_key=2026-04").data
    sheet = load_workbook(io.BytesIO(xlsx)).active
    formula_typed = [c.value for row in sheet.iter_rows() for c in row if c.data_type == "f"]
    assert not formula_typed, f"数式型セルが残っている: {formula_typed}"
    string_cells = [c.value for row in sheet.iter_rows() for c in row if isinstance(c.value, str)]
    assert not [v for v in string_cells if v and v[0] in "=+-@"], "先頭が数式記号の生セルが残っている"
    assert any(v.startswith("'=HYPERLINK") for v in string_cells)


def test_normal_csv_roundtrip_is_preserved(tmp_path):
    """通常データはエクスポート→取込で値が保持される（無害化の副作用が無い）。"""
    module, client = _build_client(tmp_path)
    module.current_user = _owner()

    entries = {
        "1": [{"id": "e1", "value": "!A!田中 一郎", "comment": "午前"}],
        "5": [{"id": "e2", "value": "!P!佐藤 花子"}],
    }
    pid = _create_with_payload(client, "新宿営業所", entries)
    csv_bytes = client.get(f"{BASE}/api/project/{pid}/export/csv?month_key=2026-04").data

    # 取込（/api/create に csv_file を渡す）。
    resp = client.post(
        f"{BASE}/api/create",
        data={"csv_file": (io.BytesIO(csv_bytes), "export.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200, resp.data
    imported = module.unwrap = resp.get_json()["project"]
    month = imported["month"]
    flat = {d: [e.get("value") for e in es] for d, es in month["entries_per_day"].items() if es}
    assert flat.get("1") == ["!A!田中 一郎"]
    assert flat.get("5") == ["!P!佐藤 花子"]
    # コメントも保持。
    day1 = month["entries_per_day"]["1"][0]
    assert day1.get("comment") == "午前"
