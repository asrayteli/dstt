"""Tests for PowerCSV enhancements: Excel-style headers, SQLite & xlsx import,
and CSV/xlsx/SQLite export endpoints."""
import io
import sqlite3
import sys
from pathlib import Path

import pytest
from flask_login import login_user

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


@pytest.fixture()
def csvtool_client():
    from app import create_app
    from app.models import User, db

    app = create_app(
        {
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SECRET_KEY": "test",
        }
    )

    with app.app_context():
        db.create_all()
        db.session.add(User(username="tester", password_hash="x", name="Tester"))
        db.session.commit()

    @app.route("/_login_test")
    def _login_test():
        login_user(User.query.filter_by(username="tester").one())
        return "ok"

    with app.test_client() as client:
        client.get("/_login_test")
        yield client


def _sqlite_bytes():
    path = Path(__file__).parent / "_tmp_test.db"
    if path.exists():
        path.unlink()
    conn = sqlite3.connect(path)
    conn.execute("CREATE TABLE people (name TEXT, age INTEGER)")
    conn.executemany("INSERT INTO people VALUES (?,?)", [("田中", 30), ("佐藤", 25)])
    conn.execute("CREATE TABLE empty_tbl (x TEXT)")
    conn.commit()
    conn.close()
    data = path.read_bytes()
    path.unlink()
    return data


def _xlsx_bytes(with_header=True):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "メイン"
    if with_header:
        ws.append(["氏名", "点数"])
    ws.append(["山田", 80])
    ws.append(["鈴木", 95])
    wb.create_sheet("別シート").append(["a", "b"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- Excel-style default headers ----------

def test_headerless_upload_uses_letter_headers(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/upload",
        data={"file": (io.BytesIO(b"1,2,3\n4,5,6\n"), "x.csv"), "has_header": "false"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["headers"] == ["A", "B", "C"]


def test_extra_columns_get_letter_headers(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/upload",
        data={"file": (io.BytesIO("名前\nまる,1,2\n".encode("utf-8")), "x.csv"), "has_header": "true"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["headers"][0] == "名前"
    assert body["headers"][1:] == ["B", "C"]


# ---------- SQLite import ----------

def test_db_tables_lists_tables(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/db/tables",
        data={"file": (io.BytesIO(_sqlite_bytes()), "d.db")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    names = {t["name"] for t in resp.get_json()["tables"]}
    assert {"people", "empty_tbl"} <= names


def test_db_load_returns_rows(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/db/load",
        data={"file": (io.BytesIO(_sqlite_bytes()), "d.db"), "table": "people"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["headers"] == ["name", "age"]
    assert body["data"] == [["田中", "30"], ["佐藤", "25"]]
    assert body["source"] == "sqlite"


def test_db_load_rejects_unknown_table(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/db/load",
        data={"file": (io.BytesIO(_sqlite_bytes()), "d.db"), "table": "people; DROP TABLE people"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400


def test_db_tables_rejects_non_sqlite(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/db/tables",
        data={"file": (io.BytesIO(b"not a database"), "d.db")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400


# ---------- xlsx import ----------

def test_xlsx_sheets_lists_sheets(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/xlsx/sheets",
        data={"file": (io.BytesIO(_xlsx_bytes()), "b.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    names = [s["name"] for s in resp.get_json()["sheets"]]
    assert names[0] == "メイン" and "別シート" in names


def test_xlsx_load_reads_sheet(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/xlsx/load",
        data={"file": (io.BytesIO(_xlsx_bytes()), "b.xlsx"), "sheet": "メイン", "has_header": "true"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["headers"] == ["氏名", "点数"]
    assert body["data"] == [["山田", "80"], ["鈴木", "95"]]


# ---------- exports ----------

def test_export_xlsx_roundtrip(csvtool_client):
    payload = {"headers": ["a", "b"], "data": [["1", "x"], ["2", "y"]], "filename": "out", "sheetName": "S"}
    resp = csvtool_client.post("/tools/csvtool/api/export/xlsx", json=payload)
    assert resp.status_code == 200
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["S"]
    assert [c.value for c in ws[1]] == ["a", "b"]
    assert [c.value for c in ws[2]] == ["1", "x"]


def test_export_db_roundtrip(csvtool_client):
    payload = {"headers": ["名前", "年齢"], "data": [["A", "1"], ["B", "2"]], "filename": "out", "tableName": "t"}
    resp = csvtool_client.post("/tools/csvtool/api/export/db", json=payload)
    assert resp.status_code == 200
    assert resp.data.startswith(b"SQLite format 3\x00")
    tmp = Path(__file__).parent / "_roundtrip.db"
    tmp.write_bytes(resp.data)
    try:
        conn = sqlite3.connect(tmp)
        rows = conn.execute('SELECT * FROM "t"').fetchall()
        cols = [d[1] for d in conn.execute('PRAGMA table_info("t")').fetchall()]
        conn.close()
    finally:
        tmp.unlink()
    assert cols == ["名前", "年齢"]
    assert rows == [("A", "1"), ("B", "2")]


def test_export_db_dedupes_and_sanitizes(csvtool_client):
    payload = {"headers": ["x", "x", ""], "data": [["1", "2", "3"]], "tableName": "9bad name!"}
    resp = csvtool_client.post("/tools/csvtool/api/export/db", json=payload)
    assert resp.status_code == 200
    tmp = Path(__file__).parent / "_dedupe.db"
    tmp.write_bytes(resp.data)
    try:
        conn = sqlite3.connect(tmp)
        tname = conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchone()[0]
        cols = [d[1] for d in conn.execute(f'PRAGMA table_info("{tname}")').fetchall()]
        conn.close()
    finally:
        tmp.unlink()
    assert len(cols) == 3 and len(set(cols)) == 3
    assert not tname[0].isdigit()


def test_csv_download_preserves_encoding(csvtool_client):
    payload = {"headers": ["名前"], "data": [["田中"]], "encoding": "shift_jis", "filename": "sjis.csv"}
    resp = csvtool_client.post("/tools/csvtool/api/download", json=payload)
    assert resp.status_code == 200
    assert "田中".encode("shift_jis") in resp.data


# ---------- regression: bug-hunt fixes ----------

def _wal_sqlite_bytes():
    """A checkpointed WAL-mode DB returns all rows via mode=ro (no immutable)."""
    path = Path(__file__).parent / "_wal_test.db"
    for p in (path, Path(str(path) + "-wal"), Path(str(path) + "-shm")):
        if p.exists():
            p.unlink()
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("CREATE TABLE w (n INTEGER)")
    conn.executemany("INSERT INTO w VALUES (?)", [(i,) for i in range(50)])
    conn.commit()
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.close()
    data = path.read_bytes()
    for p in (path, Path(str(path) + "-wal"), Path(str(path) + "-shm")):
        if p.exists():
            p.unlink()
    return data


def test_db_load_reads_wal_database(csvtool_client):
    resp = csvtool_client.post(
        "/tools/csvtool/api/db/load",
        data={"file": (io.BytesIO(_wal_sqlite_bytes()), "w.db"), "table": "w"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["totalRows"] == 50


def test_upload_rejects_oversized_file(csvtool_client, monkeypatch):
    import app.tools.csvtool as ct

    monkeypatch.setattr(ct, "MAX_UPLOAD_BYTES", 10)
    resp = csvtool_client.post(
        "/tools/csvtool/api/upload",
        data={"file": (io.BytesIO(b"a,b,c\n1,2,3\n4,5,6\n"), "big.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 413


def test_xlsx_cell_str_handles_nan_inf_and_floats():
    from app.tools.csvtool import _xlsx_cell_to_str

    assert _xlsx_cell_to_str(float("nan")) == ""
    assert _xlsx_cell_to_str(float("inf")) == ""
    assert _xlsx_cell_to_str(float("-inf")) == ""
    assert _xlsx_cell_to_str(3.0) == "3"
    assert _xlsx_cell_to_str(3.5) == "3.5"
    assert _xlsx_cell_to_str(True) == "TRUE"


def test_xlsx_load_drops_nan_cell(csvtool_client):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["v"])
    ws["A2"] = float("nan")
    ws["A3"] = 7
    buf = io.BytesIO()
    wb.save(buf)
    resp = csvtool_client.post(
        "/tools/csvtool/api/xlsx/load",
        data={"file": (io.BytesIO(buf.getvalue()), "n.xlsx"), "sheet": "Sheet", "has_header": "true"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["data"] == [[""], ["7"]]
