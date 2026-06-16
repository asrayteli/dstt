import io
import sys
from pathlib import Path
from types import SimpleNamespace

import importlib.util
from flask import Flask


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
SITE_PACKAGES = ROOT / "Lib" / "site-packages"
if SITE_PACKAGES.exists() and str(SITE_PACKAGES) not in sys.path:
    sys.path.append(str(SITE_PACKAGES))

from openpyxl import load_workbook

from app.models import SiteContractMaster, db


def _load_sat_module():
    module_path = ROOT / "app" / "tools" / "subject_analysis_tool.py"
    spec = importlib.util.spec_from_file_location("subject_analysis_tool_test_module", module_path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


def _build_client(tmp_path):
    module = _load_sat_module()
    app = Flask(
        __name__,
        root_path=str(ROOT / "app"),
        template_folder="templates",
        instance_path=str(tmp_path / "instance"),
    )
    app.secret_key = "test"
    app.config["TESTING"] = True
    app.config["LOGIN_DISABLED"] = True
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{(tmp_path / 'sat.db').as_posix()}"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(app)
    app.register_blueprint(module.subject_analysis_tool_bp)
    with app.app_context():
        db.create_all()
    return module, app.test_client()


def _user(user_id="tester01"):
    return SimpleNamespace(is_authenticated=True, username=user_id, name="Test User")


def _subject_csv_bytes(contract_type="基本請負料", subject_name="基本請負料"):
    header = [f"h{i}" for i in range(25)]
    row = [""] * 25
    row[5] = contract_type
    row[8] = "01234001"
    row[9] = "株式会社テスト"
    row[10] = "テスト現場"
    row[11] = "CODE"
    row[12] = subject_name
    row[13] = "100"
    csv_text = ",".join(header) + "\n" + ",".join(row) + "\n"
    return csv_text.encode("utf-8")


def test_sat_renderers_escape_csv_derived_html():
    chart = (ROOT / "app" / "static" / "subject_analysis_tool" / "js" / "chart_renderer.js").read_text(encoding="utf-8")
    table = (ROOT / "app" / "static" / "subject_analysis_tool" / "js" / "table_renderer.js").read_text(encoding="utf-8")
    assert "escapeText(site)" in chart
    assert "this.escapeText(data.siteName)" in table
    assert "this.escapeText(data.subjectName)" in table


def test_sat_filter_escapes_css_attribute_selectors():
    script = (ROOT / "app" / "static" / "subject_analysis_tool" / "js" / "filter_controller.js").read_text(encoding="utf-8")
    assert "escapeCssAttrValue" in script
    assert 'data-group-key="${safeGroupKey}"' in script


def test_parse_subject_data_splits_auto_sales_by_contract_type(tmp_path):
    module = _load_sat_module()
    csv_text = (
        ",".join(f"h{i}" for i in range(25))
        + "\n"
        + ",".join(
            [
                "",
                "",
                "",
                "",
                "",
                "基本請負料",
                "",
                "",
                "01234001",
                "株式会社テスト",
                "テスト現場",
                "CODE1",
                "自動車売上",
                "100",
            ]
            + [""] * 11
        )
        + "\n"
        + ",".join(
            [
                "",
                "",
                "",
                "",
                "",
                "その他請負料",
                "",
                "",
                "01234002",
                "株式会社テスト",
                "別現場",
                "CODE2",
                "自動車売上",
                "200",
            ]
            + [""] * 11
        )
        + "\n"
    )

    subject_csv_path = tmp_path / "subject.csv"
    subject_csv_path.write_text(csv_text, encoding="utf-8")
    csv_data = module.read_csv_with_encoding(subject_csv_path)
    parsed = module.parse_subject_data(csv_data)

    assert [item["subject_name"] for item in parsed] == ["基本請負料", "その他請負料"]
    assert [item["original_subject_name"] for item in parsed] == ["自動車売上", "自動車売上"]
    assert [item["is_revenue"] for item in parsed] == [True, True]
    assert [item["amounts"][0] for item in parsed] == [100.0, 200.0]


def _subject_row(contract_type="", subject_code="", subject_name="", amount="100", contract_code="01234001"):
    row = [""] * 25
    row[5] = contract_type
    row[8] = contract_code
    row[9] = "株式会社テスト"
    row[10] = "テスト現場"
    row[11] = subject_code
    row[12] = subject_name
    row[13] = amount
    return row


def _csv_from_rows(rows):
    header = [f"h{i}" for i in range(25)]
    return [header] + rows


def test_parse_subject_data_skips_sales_expense_rows():
    module = _load_sat_module()
    parsed = module.parse_subject_data(
        _csv_from_rows([
            _subject_row(subject_code="販売費", subject_name="給料", amount="500"),
            _subject_row(subject_code="CODE", subject_name="給料", amount="300"),
        ])
    )

    assert len(parsed) == 1
    assert parsed[0]["subject_name"] == "給料"
    # 原価行は符号反転されて負の値で保持される
    assert parsed[0]["amounts"][0] == -300.0
    assert parsed[0]["is_revenue"] is False


def test_parse_subject_data_passenger_sales_treated_as_revenue():
    module = _load_sat_module()
    parsed = module.parse_subject_data(
        _csv_from_rows([
            _subject_row(contract_type="基本請負料", subject_name="旅客運送売上", amount="400"),
            _subject_row(contract_type="特殊形態", subject_name="自動車売上", amount="250"),
        ])
    )

    assert [item["subject_name"] for item in parsed] == ["基本請負料", "自動車売上(特殊形態)"]
    assert all(item["is_revenue"] for item in parsed)
    # 売上行は符号反転されない
    assert [item["amounts"][0] for item in parsed] == [400.0, 250.0]


def test_read_csv_with_encoding_strips_utf8_bom(tmp_path):
    module = _load_sat_module()
    csv_path = tmp_path / "bom.csv"
    csv_path.write_bytes("﻿ヘッダ,2\nデータ,3\n".encode("utf-8"))

    csv_data = module.read_csv_with_encoding(csv_path)

    assert csv_data[0][0] == "ヘッダ"
    assert csv_data[1][0] == "データ"


def test_load_site_mapping_from_master_is_single_definition():
    source = (ROOT / "app" / "tools" / "subject_analysis_tool.py").read_text(encoding="utf-8")
    assert source.count("def load_site_mapping_from_master(") == 1


def test_subject_analysis_tool_rejects_short_required_columns(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    response = client.post(
        "/tools/subject_analysis_tool/api/upload",
        data={
            "subject_file": (io.BytesIO("a,b,c\n1,2,3\n".encode("utf-8")), "subject.csv"),
            "site_source": "file",
            "site_file": (io.BytesIO("contract,segment\n01234001,一般\n".encode("utf-8")), "site.csv"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "必須列" in response.get_json()["error"]


def test_subject_analysis_tool_db_mode_uses_site_contract_master(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    with client.application.app_context():
        db.session.add(
            SiteContractMaster(
                contract_code="01234001",
                site_row_id=1,
                site_branch_row_id=1,
                site_id="01234",
                site_branch="001",
                site_name="テスト現場",
                site_manager_id="9001",
                site_manager_name="山田 太郎",
                segment="一般",
                cloudshift_option_key="PENDING",
                is_active=True,
                source="siteplus",
            )
        )
        db.session.commit()

    response = client.post(
        "/tools/subject_analysis_tool/api/upload",
        data={
            "subject_file": (io.BytesIO(_subject_csv_bytes()), "subject.csv"),
            "site_source": "db",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["metadata"]["site_source"] == "db"
    assert payload["metadata"]["validation"]["row_count"] == 1
    assert payload["metadata"]["validation"]["unmatched_site_mapping_count"] == 0
    assert payload["data"]["site_master"]["01234001"]["site_manager_id"] == "9001"
    assert payload["data"]["site_master"]["01234001"]["segment"] == "一般"
    assert payload["data"]["site_mapping"]["01234001"] == "一般"


def _master_row(contract_code, manager_id, site_name="テスト現場", row_id=1):
    return SiteContractMaster(
        contract_code=contract_code,
        site_row_id=row_id,
        site_branch_row_id=row_id,
        site_id=contract_code[:5],
        site_branch=contract_code[5:],
        site_name=site_name,
        site_manager_id=manager_id,
        site_manager_name="担当 太郎",
        segment="一般",
        cloudshift_option_key="PENDING",
        is_active=True,
        source="siteplus",
    )


def _two_site_subject_csv_bytes():
    header = [f"h{i}" for i in range(25)]
    rows = []
    for contract_code, site_name in [("01234001", "現場A"), ("05678001", "現場B")]:
        row = [""] * 25
        row[5] = "基本請負料"
        row[8] = contract_code
        row[9] = "株式会社テスト"
        row[10] = site_name
        row[11] = "CODE"
        row[12] = "基本請負料"
        row[13] = "100"
        rows.append(row)
    csv_text = ",".join(header) + "\n" + "\n".join(",".join(row) for row in rows) + "\n"
    return csv_text.encode("utf-8")


def test_upload_manager_filter_limits_rows_to_manager_sites(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    with client.application.app_context():
        db.session.add(_master_row("01234001", "9001", "現場A", row_id=1))
        db.session.add(_master_row("05678001", "9002", "現場B", row_id=2))
        db.session.commit()

    response = client.post(
        "/tools/subject_analysis_tool/api/upload",
        data={
            "subject_file": (io.BytesIO(_two_site_subject_csv_bytes()), "subject.csv"),
            "site_source": "db",
            "manager_id": "9001",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    codes = {item["contract_code"] for item in payload["data"]["current_year"]}
    assert codes == {"01234001"}
    manager_filter = payload["metadata"]["manager_filter"]
    assert manager_filter["manager_id"] == "9001"
    assert manager_filter["matched_contract_count"] == 1
    assert manager_filter["excluded_row_count"] == 1


def test_upload_manager_filter_rejects_unknown_manager(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    with client.application.app_context():
        db.session.add(_master_row("01234001", "9001"))
        db.session.commit()

    response = client.post(
        "/tools/subject_analysis_tool/api/upload",
        data={
            "subject_file": (io.BytesIO(_subject_csv_bytes()), "subject.csv"),
            "site_source": "db",
            "manager_id": "9999",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "担当者番号" in response.get_json()["error"]


def test_upload_without_manager_id_keeps_all_rows(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    with client.application.app_context():
        db.session.add(_master_row("01234001", "9001", "現場A", row_id=1))
        db.session.add(_master_row("05678001", "9002", "現場B", row_id=2))
        db.session.commit()

    response = client.post(
        "/tools/subject_analysis_tool/api/upload",
        data={
            "subject_file": (io.BytesIO(_two_site_subject_csv_bytes()), "subject.csv"),
            "site_source": "db",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["metadata"]["manager_filter"] is None
    assert payload["metadata"]["current_year_count"] == 2


def test_managers_endpoint_lists_managers_from_master(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    with client.application.app_context():
        db.session.add(_master_row("01234001", "9001", "現場A", row_id=1))
        db.session.add(_master_row("01234002", "9001", "現場A別棟", row_id=2))
        db.session.add(_master_row("05678001", "9002", "現場B", row_id=3))
        db.session.commit()

    response = client.get("/tools/subject_analysis_tool/api/managers")

    assert response.status_code == 200
    managers = response.get_json()["managers"]
    assert [m["manager_id"] for m in managers] == ["9001", "9002"]
    assert managers[0]["site_count"] == 2
    assert managers[1]["site_count"] == 1
    assert managers[0]["manager_name"] == "担当 太郎"


def test_manager_ids_match_handles_excel_decimal_format():
    from app.site_contract_master import manager_ids_match

    assert manager_ids_match("9001.0", "9001")
    assert manager_ids_match("09001", "9001.0")
    assert not manager_ids_match("9002.0", "9001")


def test_subject_analysis_tool_page_renders_site_source_modes(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    response = client.get("/tools/subject_analysis_tool/")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'id="site-source"' in html
    assert 'id="prev-year-panel"' in html
    assert 'id="site-source-file-panel"' in html
    assert 'id="site-source-db-panel"' in html
    assert 'id="manager-id-filter"' in html
    assert 'id="upload-manager-select"' in html
    assert 'id="review-manager-filter"' in html
    assert 'id="upload-review-panel"' in html
    assert 'id="filter-preset-select"' in html
    assert 'id="sat-sidebar-toggle"' in html
    assert 'id="sat-conditions-toggle"' in html


def test_subject_analysis_tool_xlsx_export(tmp_path):
    module, client = _build_client(tmp_path)
    module.current_user = _user()

    response = client.post(
        "/tools/subject_analysis_tool/api/export/xlsx",
        json={
            "filters": {"comparisonMode": "prev_year", "months": [4]},
            "results": [
                {
                    "month": 4,
                    "currentValue": 100,
                    "comparisonValue": 80,
                    "diff": 20,
                    "diffRate": 25,
                    "hasComparison": True,
                    "isAnomaly": True,
                    "segment": "一般",
                    "item": {
                        "contract_code": "01234001",
                        "site_name": "テスト現場",
                        "corp_name": "株式会社テスト",
                        "subject_name": "基本請負料",
                        "is_revenue": True,
                    },
                }
            ],
        },
    )

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.data), data_only=True)
    assert workbook.sheetnames[:4] == ["詳細一覧", "現場別サマリー", "科目別サマリー", "条件"]
    assert workbook["詳細一覧"]["A2"].value == "01234001"
    assert workbook["詳細一覧"]["G2"].value == 100
